import ast
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from streamlit.testing.v1 import AppTest

from src.invoice_app.services.shopee_weekly_statement_service import (
    AdjustmentReconciliation,
    OrderReconciliation,
    READY_TO_COMMIT,
    StagedShopeeWeeklyStatement,
)
from src.invoice_app.parsers.shopee_weekly_statement_parser import (
    ParsedShopeeWeeklyStatement,
    SettlementIncomeRow,
)
from src.invoice_app.services.batch_service import (
    FIELD_LABELS,
    PLATFORM_ORDER_FIELDS,
)


APP_PATH = Path(__file__).resolve().parents[1] / "app.py"


def navigate(app: AppTest, page: str) -> AppTest:
    next(button for button in app.button if button.label == page).click().run(timeout=20)
    return app


def test_clear_current_batch_resets_state_before_a_fresh_batch(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [{"platform": "Lazada", "order_id": "OLD-1"}]
    app.session_state["products"] = []
    app.session_state["reviews"] = [{"order_id": "OLD-REVIEW"}]
    app.session_state["batch_id"] = "old-batch"
    app.session_state["pdf_count"] = 3
    app.session_state["uploader_version"] = 7
    app.session_state["All_order_filter"] = "OLD"
    app.session_state["All_product_filter"] = "OLD"
    app.session_state["All_platform_filter"] = "Shopee"
    app.session_state["Shopee_optional_order_columns"] = ["source_pdf"]
    app.session_state["Lazada_optional_product_columns"] = ["seller_sku"]
    app.session_state["All_export_success"] = True
    app.session_state["upload_notice"] = ("success", "old message")
    app.session_state["upload_result_summary"] = {"pdfs_processed": 3}
    app.session_state["duplicate_skipped"] = [{"order_id": "DUP-1"}]
    app.session_state["unsupported_files"] = [{"filename": "notes.pdf"}]
    app.session_state["processing_errors"] = [{"filename": "broken.pdf"}]
    app.run(timeout=20)

    app.session_state["pending_batch_discard_confirmation"] = True
    app.run(timeout=20)
    next(button for button in app.button if button.key == "confirm_discard_current_batch").click().run(timeout=20)
    cleared = app.session_state.filtered_state
    assert app.exception == []
    assert cleared["authenticated"] is True
    assert cleared["uploader_version"] == 8
    assert "pdf_uploader_8" not in cleared
    for key in (
        "orders",
        "products",
        "reviews",
        "batch_id",
        "pdf_count",
        "upload_notice",
        "upload_result_summary",
        "duplicate_skipped",
        "unsupported_files",
        "processing_errors",
    ):
        assert key not in cleared
    assert not any(key.startswith("pdf_uploader_") and key != "pdf_uploader_8" for key in cleared)
    assert not any(
        key.endswith(
            (
                "_order_filter",
                "_product_filter",
                "_platform_filter",
                "_optional_order_columns",
                "_optional_product_columns",
                "_export_success",
            )
        )
        for key in cleared
    )

    app.session_state["orders"] = [{"platform": "ZENXIN", "order_id": "NEW-1"}]
    app.session_state["products"] = []
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "fresh-batch"
    app.session_state["pdf_count"] = 1
    app.run(timeout=20)

    fresh = app.session_state.filtered_state
    assert app.exception == []
    assert fresh["batch_id"] == "fresh-batch"
    assert fresh["orders"] == [{"platform": "ZENXIN", "order_id": "NEW-1"}]
    assert fresh["pdf_count"] == 1


def test_processed_pdf_count_remains_cumulative_for_every_archived_pdf():
    tree = ast.parse(APP_PATH.read_text(encoding="utf-8"))
    process_uploads = next(
        node for node in tree.body if isinstance(node, ast.FunctionDef) and node.name == "process_uploads"
    )
    pdf_count_assignment = next(
        node
        for node in ast.walk(process_uploads)
        if isinstance(node, ast.Assign)
        and any(isinstance(target, ast.Attribute) and target.attr == "pdf_count" for target in node.targets)
    )

    value = pdf_count_assignment.value
    assert isinstance(value, ast.BinOp) and isinstance(value.op, ast.Add)
    assert isinstance(value.left, ast.Name) and value.left.id == "existing_pdf_count"
    assert isinstance(value.right, ast.Call)
    assert isinstance(value.right.func, ast.Name) and value.right.func.id == "len"
    assert isinstance(value.right.args[0], ast.Name) and value.right.args[0].id == "archived_pdfs"


def test_process_uploads_appends_each_pdf_before_processing_the_next_one():
    tree = ast.parse(APP_PATH.read_text(encoding="utf-8"))
    process_uploads = next(
        node for node in tree.body if isinstance(node, ast.FunctionDef) and node.name == "process_uploads"
    )
    pdf_loop = next(
        node
        for node in ast.walk(process_uploads)
        if isinstance(node, ast.For)
        and isinstance(node.iter, ast.Call)
        and isinstance(node.iter.func, ast.Name)
        and node.iter.func.id == "enumerate"
    )
    called_functions = {
        node.func.id
        for node in ast.walk(pdf_loop)
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name)
    }

    assert "process_pdf_file_with_outcome" in called_functions
    assert "append_batch_results_with_metadata" in called_functions


def test_platform_tabs_derive_manual_reviews_without_exposing_internal_payloads(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["reviews"] = [
        {
            "batch_id": "batch-review",
            "platform": " shopee ",
            "order_id": "SHP-R",
            "source_pdf": "shopee.pdf",
            "status": "Manual Review",
            "reason": "Income Completion Anchor Missing",
            "timestamp": "2026-08-21T10:00:00+00:00",
            "order_payload": {"delivery_fee": "4.90", "payment_status": "Released"},
            "product_payloads": [
                {
                    "platform": "Shopee",
                    "order_id": "SHP-R",
                    "product_name": "Review product",
                    "unit_price": "10.00",
                    "seller_sku": "SKU-R",
                    "quantity": 1,
                }
            ],
        },
        {
            "batch_id": "batch-review",
            "platform": "Lazada",
            "order_id": "LZD-R",
            "source_pdf": "lazada.pdf",
            "status": "Manual Review",
            "reason": "Product Count Mismatch",
            "timestamp": "2026-08-21T10:01:00+00:00",
        },
        {
            "batch_id": "batch-review",
            "platform": "ZENXIN",
            "order_id": "ZNX-R",
            "source_pdf": "zenxin.pdf",
            "status": "Manual Review",
            "reason": "Product Count Mismatch",
            "timestamp": "2026-08-21T10:02:00+00:00",
        },
    ]
    app.session_state["orders"] = []
    app.session_state["products"] = []
    app.session_state["batch_id"] = "batch-review"
    app.session_state["pdf_count"] = 3
    app.session_state["navigation"] = "Shopee"

    app.run(timeout=20)

    for platform_name in ("Shopee", "Lazada", "ZENXIN"):
        if platform_name != "Shopee":
            navigate(app, platform_name)
        assert app.exception == []
        assert "Manual Review" in {element.value for element in app.subheader}
        assert ("Manual Review", "1") in {
            (metric.label, metric.value) for metric in app.metric
        }
        assert app.expander == []
        for dataframe in app.dataframe:
            assert "order_payload" not in dataframe.value.columns
            assert "product_payloads" not in dataframe.value.columns
        if platform_name == "Shopee":
            review_table = next(
                dataframe.value
                for dataframe in app.dataframe
                if "Payment Status" in dataframe.value.columns
            )
            assert review_table["Payment Status"].tolist() == ["Released"]
        else:
            assert all("Payment Status" not in dataframe.value.columns for dataframe in app.dataframe)



def test_all_tab_separates_incomplete_product_rows_without_exporting_them(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = []
    app.session_state["products"] = []
    app.session_state["reviews"] = [
        {
            "batch_id": "batch-all-review",
            "platform": "Lazada",
            "order_id": "LZD-R",
            "source_pdf": "lazada.pdf",
            "status": "Manual Review",
            "reason": "Financial Reconciliation Failed",
            "timestamp": "2026-08-21T10:00:00+00:00",
            "order_payload": {"delivery_fee": "N/A"},
            "product_payloads": [
                {
                    "platform": "Lazada",
                    "order_id": "LZD-R",
                    "product_name": "Complete product",
                    "unit_price": "12.00",
                    "paid_price": "9.00",
                    "seller_sku": "SKU-A",
                    "quantity": 1,
                },
                {
                    "platform": "Lazada",
                    "order_id": "LZD-R",
                    "product_name": "Missing price product",
                    "unit_price": "N/A",
                    "seller_sku": "SKU-B",
                    "quantity": 1,
                },
            ],
        }
    ]
    app.session_state["batch_id"] = "batch-all-review"
    app.session_state["pdf_count"] = 1
    app.session_state["navigation"] = "Cross Platform Summary"

    app.run(timeout=20)

    assert app.exception == []
    assert app.expander == []
    assert {"Product Summary", "Filters", "All Products", "All Manual Review"} <= {
        element.value for element in app.subheader
    }
    assert {"From Date", "To Date"} <= {date_input.label for date_input in app.date_input}
    assert all(date_input.label != "Order Created Date Range" for date_input in app.date_input)
    platform_filter = next(
        selectbox for selectbox in app.selectbox if selectbox.label == "Platform"
    )
    assert platform_filter.options == ["All", "Shopee", "Lazada", "ZENXIN"]
    frames_by_columns = {tuple(frame.value.columns): frame.value for frame in app.dataframe}
    all_products = frames_by_columns[
        ("Order Created Date",
        "Product Name",
        "Product Price",
        "Seller SKU #",
        "Qty",
        "Platform")
    ]
    assert all_products["Product Name"].tolist() == ["Complete product"]
    assert all_products["Product Price"].tolist() == ["12.00"]
    assert "Data Status" not in all_products.columns
    product_summary = frames_by_columns[
        ("Seller SKU", "Product Name", "Unit Selling Price", "Total Quantity", "Total Selling Price", "Total Discount Given")
    ]
    assert product_summary["Seller SKU"].tolist() == ["SKU-A"]
    assert product_summary["Total Quantity"].tolist() == [1]
    assert product_summary["Total Selling Price"].tolist() == [9.0]
    assert {"Export All Products", "Export Product Summary"} <= {
        button.label for button in app.get("download_button")
    }
    all_review = frames_by_columns[
        (
            "Product Name",
            "Product Price",
            "Seller SKU #",
            "Qty",
            "Delivery Fee",
            "Platform",
            "All Review Reason",
        )
    ]
    assert all_review["Product Name"].tolist() == ["Missing price product"]
    assert all_review["All Review Reason"].tolist() == ["Missing or invalid Product Price"]


def test_platform_tabs_hide_manual_review_section_when_that_platform_has_none(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = []
    app.session_state["products"] = []
    app.session_state["reviews"] = [
        {
            "batch_id": "batch-shopee-only",
            "platform": "Shopee",
            "order_id": "SHP-R",
            "source_pdf": "shopee.pdf",
            "status": "Manual Review",
            "reason": "Income Completion Anchor Missing",
            "timestamp": "2026-08-21T10:00:00+00:00",
        }
    ]
    app.session_state["batch_id"] = "batch-shopee-only"
    app.session_state["pdf_count"] = 1
    app.session_state["navigation"] = "Shopee"

    app.run(timeout=20)

    assert app.exception == []
    assert "Manual Review" in {element.value for element in app.subheader}
    assert ("Manual Review", "1") in {
        (metric.label, metric.value) for metric in app.metric
    }

    for platform_name in ("Lazada", "ZENXIN"):
        navigate(app, platform_name)
        assert app.exception == []
        assert "Manual Review" not in {element.value for element in app.subheader}
        assert ("Manual Review", "0") in {
            (metric.label, metric.value) for metric in app.metric
        }


def test_lazada_preview_dates_are_typed_for_chronological_sorting(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [
        {
            "platform": "Lazada",
            "order_id": "LZD-DATE-1",
            "order_date": "28 02 2026",
            "invoice_date": "27 02 2026",
        },
        {
            "platform": "Lazada",
            "order_id": "LZD-DATE-2",
            "order_date": "01 03 2026",
            "invoice_date": "02 03 2026",
        },
    ]
    app.session_state["products"] = []
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "batch-lazada-dates"
    app.session_state["pdf_count"] = 2
    app.session_state["navigation"] = "Lazada"

    app.run(timeout=20)

    assert app.exception == []
    order_table = next(
        dataframe.value
        for dataframe in app.dataframe
        if {"Order Date", "Invoice Date"} <= set(dataframe.value.columns)
    )
    assert order_table["Order Date"].dtype.kind == "M"
    assert order_table["Invoice Date"].dtype.kind == "M"

def test_zenxin_preview_invoice_date_is_typed_without_lazada_format_coercion(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [
        {
            "platform": "ZENXIN",
            "order_id": "10123",
            "invoice_date": "31/03/2026",
        }
    ]
    app.session_state["products"] = []
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "batch-zenxin-date"
    app.session_state["pdf_count"] = 1
    app.session_state["navigation"] = "ZENXIN"

    app.run(timeout=20)

    assert app.exception == []
    order_table = next(
        dataframe.value
        for dataframe in app.dataframe
        if "Invoice Date" in dataframe.value.columns
    )
    assert order_table["Invoice Date"].dtype.kind == "M"
    assert order_table["Invoice Date"].iloc[0].strftime("%d/%m/%Y") == "31/03/2026"

def test_platform_order_defaults_are_compact_for_shopee_and_lazada_only(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [
        {"platform": "Shopee", "order_id": "SHP-1", "payment_status": "Pending"},
        {"platform": "Lazada", "order_id": "LZD-1"},
        {"platform": "ZENXIN", "order_id": "ZNX-1"},
    ]
    app.session_state["products"] = []
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "batch-default-columns"
    app.session_state["pdf_count"] = 3
    app.session_state["navigation"] = "Shopee"

    app.run(timeout=20)
    assert app.exception == []
    shopee_order_table = next(
        dataframe.value
        for dataframe in app.dataframe
        if "Payment Status" in dataframe.value.columns
    )
    assert shopee_order_table["Payment Status"].tolist() == ["Pending"]
    navigate(app, "Lazada")
    assert app.exception == []
    navigate(app, "ZENXIN")

    assert app.exception == []
    state = app.session_state.filtered_state
    assert state["Shopee_optional_order_columns"] == [
        "order_status",
        "payment_status",
        "order_created_date",
        "fund_transfer_date",
        "order_income",
        "merchandise_subtotal",
    ]
    assert state["Lazada_optional_order_columns"] == [
        "invoice_number",
        "order_date",
        "invoice_date",
        "payment_method",
        "subtotal",
        "net_paid",
        "source_pdf",
        "status",
    ]
    assert state["ZENXIN_optional_order_columns"] == [
        column
        for column in PLATFORM_ORDER_FIELDS["ZENXIN"]
        if column not in {"platform", "order_id"}
    ]


def test_shopee_order_table_projects_missing_created_date_from_order_id(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [
        {"platform": "Shopee", "order_id": "260828J247W9SW", "order_created_date": "N/A"}
    ]
    app.session_state["products"] = []
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "batch-order-date-projection"
    app.session_state["pdf_count"] = 1
    app.session_state["navigation"] = "Shopee"

    app.run(timeout=20)

    assert app.exception == []
    order_table = next(
        dataframe.value
        for dataframe in app.dataframe
        if "Order Created Date" in dataframe.value.columns
    )
    assert order_table["Order Created Date"].iloc[0].date() == date(2026, 8, 28)


def test_platform_tables_keep_all_fields_available_in_native_toolbar(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [
        {
            "platform": "Lazada",
            "order_id": "LZD-1",
            "voucher_applied": "1.20",
        }
    ]
    app.session_state["products"] = []
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "batch-full-table"
    app.session_state["pdf_count"] = 1
    app.session_state["navigation"] = "Lazada"

    app.run(timeout=20)

    assert app.exception == []
    order_table = next(
        dataframe for dataframe in app.dataframe if "Order ID" in dataframe.value.columns
    )
    assert order_table.value.columns.tolist() == [
        FIELD_LABELS.get(column, column)
        for column in PLATFORM_ORDER_FIELDS["Lazada"]
    ]
    assert "Voucher Applied" in order_table.value.columns
    assert app.session_state.filtered_state["Lazada_optional_order_columns"] == [
        "invoice_number",
        "order_date",
        "invoice_date",
        "payment_method",
        "subtotal",
        "net_paid",
        "source_pdf",
        "status",
    ]


def test_upload_summary_is_action_scoped_and_skipped_items_stay_out_of_manual_review(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["batch_id"] = "batch-summary"
    app.session_state["pdf_count"] = 15
    app.session_state["orders"] = [
        {"platform": "Shopee", "order_id": "ORD-A", "delivery_fee": "2.00"}
    ]
    app.session_state["products"] = [
        {
            "platform": "Shopee",
            "order_id": "ORD-A",
            "product_name": "Accepted product",
            "unit_price": "10.00",
            "seller_sku": "SKU-A",
            "quantity": 1,
        }
    ]
    app.session_state["reviews"] = [
        {
            "batch_id": "batch-summary",
            "platform": "Shopee",
            "order_id": "ORD-R",
            "source_pdf": "review.pdf",
            "status": "Manual Review",
            "reason": "Financial Reconciliation Failed",
            "order_payload": {"delivery_fee": "1.00"},
            "product_payloads": [
                {
                    "platform": "Shopee",
                    "order_id": "ORD-R",
                    "product_name": "Review product",
                    "unit_price": "12.00",
                    "seller_sku": "SKU-R",
                    "quantity": 1,
                }
            ],
        },
        {
            "batch_id": "batch-summary",
            "platform": "Shopee",
            "order_id": "ORD-D",
            "source_pdf": "duplicate.pdf",
            "status": "Duplicate Skipped",
            "reason": "Duplicate Order",
            "order_payload": {"delivery_fee": "99.00"},
            "product_payloads": [
                {
                    "platform": "Shopee",
                    "order_id": "ORD-D",
                    "product_name": "Duplicate product",
                    "unit_price": "99.00",
                    "seller_sku": "SKU-D",
                    "quantity": 9,
                }
            ],
        },
    ]
    app.session_state["upload_result_summary"] = {
        "pdfs_processed": 5,
        "orders_imported": 3,
        "manual_reviews": 1,
        "duplicate_orders": 1,
        "unsupported_files": 1,
        "processing_errors": 0,
    }
    app.session_state["duplicate_skipped"] = [
        {
            "platform": "Shopee",
            "order_id": "ORD-D",
            "source_pdf": "duplicate.pdf",
            "message": "Already exists in current batch.",
        }
    ]
    app.session_state["unsupported_files"] = [
        {
            "filename": "lecture_notes.pdf",
            "source_pdf": "lecture_notes.pdf",
            "status": "Unsupported",
            "message": "Not recognized as a supported invoice.",
        }
    ]
    app.session_state["processing_errors"] = []

    app.session_state["navigation"] = "Data Import"
    app.run(timeout=20)

    assert app.exception == []
    assert {
        "Data Import",
        "Dashboard",
        "Cross Platform Summary",
        "Shopee",
        "Lazada",
        "ZENXIN",
   } <= {button.label for button in app.button}
    metrics = {(metric.label, metric.value) for metric in app.metric}
    assert {
        ("PDFs", "15"),
        ("Orders", "1"),
        ("Products", "1"),
        ("Manual Review", "1"),
    } <= metrics
    assert "Duplicate skipped: 1 · Unsupported: 1" in {caption.value for caption in app.caption}
    assert "batch-summary" not in {caption.value for caption in app.caption}
    assert {"Discard current batch", "Logout"} <= {
        button.label for button in app.button
    }
    assert "Export Batch Excel" not in {
        button.label for button in app.get("download_button")
    }
    assert app.toggle == []
    assert app.get("badge") == []
    assert "View skipped items" in {expander.label for expander in app.expander}
    assert {"Validate", "Result Summary", "Available recovery actions"} <= {
        element.value for element in app.subheader
    }
    assert len(app.dataframe) == 0

    navigate(app, "Shopee")

    assert app.exception == []
    assert {"Search and Filters", "Order Level", "Product Level", "Manual Review"} <= {
        element.value for element in app.subheader
    }
    assert app.expander == []
    assert len(app.dataframe) == 3
    assert {"Export full Shopee batch", "Export current filtered view"} <= {
        button.label for button in app.get("download_button")
    }


def test_platform_export_keeps_full_batch_separate_from_filtered_view(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [
        {"platform": "Shopee", "order_id": "SHP-1", "order_income": "10.00"},
        {"platform": "Shopee", "order_id": "SHP-2", "order_income": "20.00"},
    ]
    app.session_state["products"] = [
        {
            "platform": "Shopee",
            "order_id": "SHP-1",
            "product_name": "First product",
            "seller_sku": "SKU-1",
            "quantity": 1,
            "line_subtotal": "10.00",
        },
        {
            "platform": "Shopee",
            "order_id": "SHP-2",
            "product_name": "Second product",
            "seller_sku": "SKU-2",
            "quantity": 2,
            "line_subtotal": "20.00",
        },
    ]
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "batch-export-scopes"
    app.session_state["pdf_count"] = 2
    app.session_state["navigation"] = "Shopee"

    app.run(timeout=20)
    next(element for element in app.text_input if element.label == "Order ID").set_value("SHP-1").run(timeout=20)

    assert app.exception == []
    assert {"Export full Shopee batch", "Export current filtered view"} <= {
        button.label for button in app.get("download_button")
    }

    full_workbook = load_workbook(tmp_path / "exports" / "batch-export-scopes-shopee-full-batch-report.xlsx")
    filtered_workbook = load_workbook(tmp_path / "exports" / "batch-export-scopes-shopee-filtered-view-report.xlsx")
    assert full_workbook["Orders"].max_row == 5
    assert full_workbook["Products"].max_row == 5
    assert filtered_workbook["Orders"].max_row == 4
    assert filtered_workbook["Products"].max_row == 4
    assert dict(full_workbook["Summary"].iter_rows(min_row=4, values_only=True))["Orders"] == 2
    assert dict(filtered_workbook["Summary"].iter_rows(min_row=4, values_only=True))["Orders"] == 1

def test_cross_platform_from_to_filters_share_the_same_detail_and_summary_population(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [
        {"platform": "Shopee", "order_id": "SHP-1", "order_created_date": "05/08/2026"},
        {"platform": "Lazada", "order_id": "LZD-1", "order_date": "15 08 2026"},
    ]
    app.session_state["products"] = [
        {
            "platform": "Shopee",
            "order_id": "SHP-1",
            "product_name": "Early product",
            "seller_sku": "SKU-EARLY",
            "unit_price": "10.00",
            "quantity": 1,
            "line_subtotal": "10.00",
        },
        {
            "platform": "Lazada",
            "order_id": "LZD-1",
            "product_name": "Later product",
            "seller_sku": "SKU-LATER",
            "unit_price": "20.00",
            "quantity": 1,
            "paid_price": "20.00",
        },
    ]
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "batch-cross-platform-filter"
    app.session_state["pdf_count"] = 2
    app.session_state["navigation"] = "Cross Platform Summary"

    app.run(timeout=20)

    from_date = next(element for element in app.date_input if element.label == "From Date")
    from_date.set_value(date(2026, 8, 8)).run(timeout=20)
    frames_by_columns = {tuple(frame.value.columns): frame.value for frame in app.dataframe}
    detail = frames_by_columns[
        ("Order Created Date", "Product Name", "Product Price", "Seller SKU #", "Qty", "Platform")
    ]
    summary = frames_by_columns[("Seller SKU", "Product Name", "Unit Selling Price", "Total Quantity", "Total Selling Price", "Total Discount Given")]
    assert detail["Product Name"].tolist() == ["Later product"]
    assert detail["Order Created Date"].iloc[0].date() == date(2026, 8, 15)
    assert summary["Seller SKU"].tolist() == ["SKU-LATER"]

    next(element for element in app.date_input if element.label == "From Date").set_value(None).run(timeout=20)
    next(element for element in app.date_input if element.label == "To Date").set_value(date(2026, 8, 8)).run(timeout=20)
    frames_by_columns = {tuple(frame.value.columns): frame.value for frame in app.dataframe}
    detail = frames_by_columns[
        ("Order Created Date", "Product Name", "Product Price", "Seller SKU #", "Qty", "Platform")
    ]
    summary = frames_by_columns[("Seller SKU", "Product Name", "Unit Selling Price", "Total Quantity", "Total Selling Price", "Total Discount Given")]
    assert detail["Product Name"].tolist() == ["Early product"]
    assert summary["Seller SKU"].tolist() == ["SKU-EARLY"]

    next(element for element in app.date_input if element.label == "From Date").set_value(date(2026, 8, 15)).run(timeout=20)
    assert any(error.value == "From Date must be on or before To Date." for error in app.error)
    frames_by_columns = {tuple(frame.value.columns): frame.value for frame in app.dataframe}
    detail = frames_by_columns[
        ("Order Created Date", "Product Name", "Product Price", "Seller SKU #", "Qty", "Platform")
    ]
    assert detail["Product Name"].tolist() == ["Early product", "Later product"]


def test_cross_platform_date_filter_with_no_matching_rows_shows_empty_state_without_error(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["orders"] = [
        {"platform": "Shopee", "order_id": "SHP-1", "order_created_date": "05/08/2026"}
    ]
    app.session_state["products"] = [
        {
            "platform": "Shopee",
            "order_id": "SHP-1",
            "product_name": "Only product",
            "seller_sku": "SKU-ONLY",
            "unit_price": "10.00",
            "quantity": 1,
            "line_subtotal": "10.00",
        }
    ]
    app.session_state["reviews"] = []
    app.session_state["batch_id"] = "batch-cross-platform-empty"
    app.session_state["pdf_count"] = 1
    app.session_state["navigation"] = "Cross Platform Summary"

    app.run(timeout=20)
    next(element for element in app.date_input if element.label == "From Date").set_value(
        date(2026, 8, 6)
    ).run(timeout=20)

    assert app.exception == []
    assert "No products match the current filters." in {caption.value for caption in app.caption}
    assert any(button.label == "Export All Products" and button.disabled for button in app.button)


def _weekly_stage_for_ui() -> StagedShopeeWeeklyStatement:
    order_row = SettlementIncomeRow(
        sequence_no="1",
        view_by="Order",
        order_id="SHP-1",
        product_id="PID-1",
        product_name="Statement product",
        order_creation_date=date(2026, 8, 1),
        payout_completed_date=date(2026, 8, 7),
        release_channel="Seller Wallet",
        order_type="Normal",
        total_released_amount=Decimal("12.34"),
        financial_components={},
        source_values={},
        source_row_number=2,
    )
    sku_row = SettlementIncomeRow(
        sequence_no="2",
        view_by="Sku",
        order_id="SHP-1",
        product_id="PID-1",
        product_name="Statement product",
        order_creation_date=date(2026, 8, 1),
        payout_completed_date=date(2026, 8, 7),
        release_channel="Seller Wallet",
        order_type="Normal",
        total_released_amount=Decimal("12.34"),
        financial_components={},
        source_values={},
        source_row_number=3,
    )
    statement = ParsedShopeeWeeklyStatement(
        source_filename="Income.released.my.xlsx",
        file_hash="test-hash",
        statement_period_from=date(2026, 8, 1),
        statement_period_to=date(2026, 8, 7),
        summary_total_released=Decimal("12.34"),
        adjustment_control_total=Decimal("-1.00"),
        adjustment_footer_total=Decimal("-1.00"),
        income_rows=(order_row, sku_row),
        service_fee_details=(),
        shipping_fee_discrepancies=(),
        adjustments=(),
        source_value_issues=(),
        dimension_fallback_sheets=(),
    )
    return StagedShopeeWeeklyStatement(
        result=READY_TO_COMMIT,
        source_filename=statement.source_filename,
        file_hash=statement.file_hash,
        statement=statement,
        validation_issues=(),
        review_reasons=(),
        rejection_reasons=(),
        duplicate_status=None,
        order_reconciliations=(
            OrderReconciliation(
                order_id="SHP-1",
                status="Different",
                released_amount=Decimal("12.34"),
                order_income=Decimal("10.00"),
                difference=Decimal("2.34"),
            ),
        ),
        adjustment_reconciliations=(
            AdjustmentReconciliation(
                linked_order_id="SHP-OLD",
                status="Unmatched Adjustment",
                adjustment_amount=Decimal("-1.00"),
            ),
        ),
    )


def test_data_import_wizard_selects_weekly_statement_before_upload(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["navigation"] = "Data Import"

    app.run(timeout=20)

    assert app.exception == []
    assert "Data Import" in {title.value for title in app.title}
    assert "Step 1 of 5 — Select Source" in {element.text for element in app.get("progress")}
    source_type = next(radio for radio in app.radio if radio.label == "Data type")
    assert source_type.options == ["Platform Orders", "Shopee Weekly Statement"]
    source_type.set_value("Shopee Weekly Statement").run(timeout=20)
    next(button for button in app.button if button.label == "Continue to upload").click().run(timeout=20)

    assert app.exception == []
    assert app.session_state.filtered_state["import_source_type"] == "Shopee Weekly Statement"
    assert app.session_state.filtered_state["data_import_step"] == 2
    assert any(element.label == "Shopee Weekly Statement (.xlsx)" for element in app.file_uploader)
    assert {"Shipment Confirmation", "Historical Import"}.isdisjoint(
        {element.value for element in app.subheader}
    )


def test_data_import_prevents_second_source_for_an_active_batch(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["navigation"] = "Data Import"
    app.session_state["batch_id"] = "active-platform-batch"
    app.session_state["import_source_type"] = "Platform Orders"
    app.session_state["data_import_step"] = 1

    app.run(timeout=20)

    assert app.exception == []
    assert "Continue current batch" in {element.value for element in app.subheader}
    assert not any(radio.label == "Data type" for radio in app.radio)
    assert {"Continue", "Discard current batch"} <= {
        button.label for button in app.button
    }


def test_weekly_statement_reconciliation_uses_existing_staged_counts(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    app = AppTest.from_file(str(APP_PATH))
    app.session_state["authenticated"] = True
    app.session_state["navigation"] = "Data Import"
    app.session_state["batch_id"] = "weekly-batch"
    app.session_state["import_source_type"] = "Shopee Weekly Statement"
    app.session_state["data_import_step"] = 4
    app.session_state["weekly_statement_stage"] = _weekly_stage_for_ui()

    app.run(timeout=20)

    assert app.exception == []
    metrics = {(metric.label, metric.value) for metric in app.metric}
    assert {
        ("Matched", "0"),
        ("Different", "1"),
        ("Estimated Only", "0"),
        ("Unmatched Orders", "0"),
        ("Unmatched Adjustments", "1"),
        ("Shipping exceptions", "0"),
    } <= metrics
    assert "Representative exceptions" in {caption.value for caption in app.caption}
    assert "These results are shown for review and do not change the source outcome." in {
        caption.value for caption in app.caption
    }
