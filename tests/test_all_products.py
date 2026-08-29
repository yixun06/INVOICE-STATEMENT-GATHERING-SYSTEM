from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from src.invoice_app.services.all_products import (
    ALL_PRODUCT_COLUMNS,
    ALL_PRODUCT_DISPLAY_COLUMNS,
    ALL_PRODUCT_DISPLAY_FIELD_LABELS,

    ALL_PRODUCT_FIELD_LABELS,
    build_all_product_rows,
    build_all_product_views,
    build_cross_platform_product_rows,
    filter_cross_platform_product_rows,
    summarize_cross_platform_products,
)
from src.invoice_app.services.exporter import export_all_products_report
from src.invoice_app.services.product_price_master import ProductPriceMaster


def test_all_product_view_reuses_accepted_product_and_order_fields_without_status():
    rows = build_all_product_rows(
        [
            {
                "platform": "Shopee",
                "order_id": "SHP-1",
                "delivery_fee": "3.20",
            }
        ],
        [
            {
                "platform": "Shopee",
                "order_id": "SHP-1",
                "product_name": "Shopee product",
                "unit_price": "12.50",
                "seller_sku": "SHP-SKU",
                "quantity": 2,
            }
        ],
    )

    assert rows == [
        {
            "product_name": "Shopee product",
            "unit_price": "12.50",
            "seller_sku": "SHP-SKU",
            "quantity": 2,
            "delivery_fee": "3.20",
            "platform": "Shopee",
            "order_id": "SHP-1",
        }
    ]


def test_all_product_display_excludes_delivery_fee_and_order_date():
    assert "delivery_fee" not in ALL_PRODUCT_DISPLAY_COLUMNS
    assert "order_date" not in ALL_PRODUCT_DISPLAY_COLUMNS
    assert ALL_PRODUCT_DISPLAY_FIELD_LABELS == {
        "product_name": "Product Name",
        "unit_price": "Product Price",
        "seller_sku": "Seller SKU #",
        "quantity": "Qty",
        "platform": "Platform",
    }



def test_all_product_view_uses_canonical_delivery_join_and_lazada_price():
    rows = build_all_product_rows(
        [{"platform": "Lazada", "order_id": " LZD-1 ", "delivery_fee": "4.50"}],
        [
            {
                "platform": "lazada",
                "order_id": "LZD-1",
                "product_name": "First Lazada product",
                "unit_price": "12.00",
                "paid_price": "9.00",
                "seller_sku": "LZD-1A",
                "quantity": 1,
            },
            {
                "platform": "LAZADA",
                "order_id": " LZD-1",
                "product_name": "Second Lazada product",
                "unit_price": "20.00",
                "paid_price": "15.00",
                "seller_sku": "LZD-1B",
                "quantity": 2,
            },
        ],
    )

    assert [row["delivery_fee"] for row in rows] == ["4.50", "4.50"]
    assert [row["unit_price"] for row in rows] == ["12.00", "20.00"]
    assert all("paid_price" not in row and "data_status" not in row for row in rows)


def test_manual_review_rows_are_classified_individually_and_keep_delivery_na():
    normal_rows, review_rows = build_all_product_views(
        [],
        [],
        [
            {
                "platform": "Shopee",
                "order_id": "SHP-R",
                "order_payload": {"delivery_fee": "N/A"},
                "product_payloads": [
                    {
                        "platform": "Shopee",
                        "order_id": "SHP-R",
                        "product_name": "Complete A",
                        "unit_price": "10.00",
                        "seller_sku": "S-A",
                        "quantity": 1,
                    },
                    {
                        "platform": "Shopee",
                        "order_id": "SHP-R",
                        "product_name": "Complete zero-price B",
                        "unit_price": "0.00",
                        "seller_sku": "S-B",
                        "quantity": 2,
                    },
                    {
                        "platform": "Shopee",
                        "order_id": "SHP-R",
                        "product_name": "Missing price C",
                        "unit_price": "N/A",
                        "seller_sku": "S-C",
                        "quantity": 1,
                    },
                ],
            }
        ],
    )

    assert [row["product_name"] for row in normal_rows] == ["Complete A", "Complete zero-price B"]
    assert [row["delivery_fee"] for row in normal_rows] == ["N/A", "N/A"]
    assert review_rows[0]["product_name"] == "Missing price C"
    assert review_rows[0]["all_review_reason"] == "Missing or invalid Product Price"


def test_manual_review_rows_with_missing_name_qty_or_platform_stay_in_all_review():
    normal_rows, review_rows = build_all_product_views(
        [],
        [],
        [
            {
                "platform": "Unknown",
                "order_id": "UNK-1",
                "order_payload": {"delivery_fee": "3.00"},
                "product_payloads": [
                    {
                        "platform": "Unknown",
                        "order_id": "UNK-1",
                        "product_name": "Missing platform",
                        "unit_price": "10.00",
                        "quantity": 1,
                    },
                    {
                        "platform": "Unknown",
                        "order_id": "UNK-1",
                        "product_name": "Missing quantity",
                        "unit_price": "10.00",
                        "quantity": "N/A",
                    },
                    {
                        "platform": "Unknown",
                        "order_id": "UNK-1",
                        "product_name": "",
                        "unit_price": "10.00",
                        "quantity": 1,
                    },
                ],
            }
        ],
    )

    assert normal_rows == []
    assert [row["all_review_reason"] for row in review_rows] == [
        "Missing or invalid Platform",
        "Missing or invalid Qty; Missing or invalid Platform",
        "Missing Product Name; Missing or invalid Platform",
    ]


def test_all_product_view_prefers_accepted_identity_over_duplicate_review_in_both_views():
    normal_rows, review_rows = build_all_product_views(
        [{"platform": "Shopee", "order_id": "ORD-1", "delivery_fee": "2.00"}],
        [
            {
                "platform": "Shopee",
                "order_id": "ORD-1",
                "product_name": "Accepted product",
                "unit_price": "10.00",
                "seller_sku": "SKU-A",
                "quantity": 1,
            }
        ],
        [
            {
                "platform": " shopee ",
                "order_id": " ORD-1 ",
                "order_payload": {"delivery_fee": "99.00"},
                "product_payloads": [
                    {
                        "platform": "SHOPEE",
                        "order_id": "ORD-1",
                        "product_name": "Duplicate product",
                        "unit_price": "N/A",
                        "seller_sku": "SKU-D",
                        "quantity": 9,
                    }
                ],
            }
        ],
    )

    assert [row["product_name"] for row in normal_rows] == ["Accepted product"]
    assert review_rows == []


def test_all_product_export_uses_requested_labels_and_numeric_formats(tmp_path):
    destination = tmp_path / "all-products.xlsx"
    export_all_products_report(
        destination=destination,
        products=[
            {
                "product_name": "All product",
                "unit_price": "12.50",
                "seller_sku": "000123",
                "quantity": 2,
                "delivery_fee": "3.20",
                "platform": "Shopee",
            }
        ],
        product_columns=ALL_PRODUCT_COLUMNS,
        column_labels=ALL_PRODUCT_FIELD_LABELS,
    )

    workbook = load_workbook(destination)
    assert workbook.sheetnames == ["Products"]
    products = workbook["Products"]
    assert products["A1"].value == "PRODUCT DETAILS - ALL"
    assert products.sheet_view.showGridLines is True
    assert products.sheet_view.zoomScale == 100
    assert not products.merged_cells.ranges
    assert products["A1"].font.name == "Calibri"
    assert products["A1"].font.color.rgb == "001F4E78"
    assert products.freeze_panes == "A4"
    assert products["A3"].fill.fgColor.rgb == "001F4E78"
    assert products["A3"].alignment.horizontal == "center"
    assert products["A4"].font.name == "Calibri"
    assert products["A4"].border.left.color.rgb == "00000000"
    assert [cell.value for cell in products[3]] == [
        ALL_PRODUCT_FIELD_LABELS[column] for column in ALL_PRODUCT_COLUMNS
    ]
    assert "Data Status" not in {cell.value for cell in products[3]}
    assert "Delivery Fee" not in {cell.value for cell in products[3]}
    assert "Order Date" not in {cell.value for cell in products[3]}
    assert products["B4"].value == 12.5
    assert products["B4"].number_format == '#,##0.00;[Red]-#,##0.00'
    assert products["C4"].value == "000123"
    assert products["C4"].number_format == "@"
    assert products["D4"].value == 2
    assert products["D4"].number_format == "#,##0"
    assert products["E4"].value == "Shopee"


def test_all_views_ignore_duplicate_and_unsupported_payloads():
    normal_rows, review_rows = build_all_product_views(
        [],
        [],
        [
            {
                "platform": "Shopee",
                "order_id": "DUP-1",
                "status": "Duplicate Skipped",
                "reason": "Duplicate Order",
                "order_payload": {"delivery_fee": "9.00"},
                "product_payloads": [
                    {
                        "platform": "Shopee",
                        "order_id": "DUP-1",
                        "product_name": "Duplicate product",
                        "unit_price": "9.00",
                        "quantity": 9,
                    }
                ],
            },
            {
                "platform": "Unknown",
                "order_id": "N/A",
                "status": "Unsupported",
                "product_payloads": [
                    {
                        "platform": "Unknown",
                        "product_name": "Unsupported product",
                        "unit_price": "1.00",
                        "quantity": 1,
                    }
                ],
            },
        ],
    )

    assert normal_rows == []
    assert review_rows == []


def test_cross_platform_summary_uses_dates_and_price_identity_rows():
    master = ProductPriceMaster.from_rows([
        {"seller_sku": "SKU-SHARED", "product_name": "Shared SKU first name", "unit_selling_price": "10.00"},
        {"seller_sku": "SKU-SHARED", "product_name": "Different source label", "unit_selling_price": "10.00"},
        {"seller_sku": "SKU-ZENXIN", "product_name": "ZENXIN product", "unit_selling_price": "11.90"},
    ])
    orders = [
        {"platform": "Shopee", "order_id": "SHP-1", "order_created_date": "07/08/2026 12:04"},
        {"platform": "Lazada", "order_id": "LZD-1", "order_date": "08 08 2026"},
        {"platform": "ZENXIN", "order_id": "ZNX-1", "invoice_date": "09/08/2026"},
    ]
    products = [
        {"platform": "Shopee", "order_id": "SHP-1", "product_name": "Shared SKU first name", "seller_sku": "SKU-SHARED", "unit_price": "10.00", "line_subtotal": "20.00", "source_line_subtotal": "20.00", "quantity": 2},
        {"platform": "Lazada", "order_id": "LZD-1", "product_name": "Different source label", "seller_sku": "SKU-SHARED", "unit_price": "12.00", "paid_price": "9.00", "quantity": 1},
        {"platform": "ZENXIN", "order_id": "ZNX-1", "product_name": "ZENXIN product", "seller_sku": "SKU-ZENXIN", "unit_price": "11.90", "line_total_inc_tax": "35.70", "quantity": 3},
    ]

    reporting_rows = build_cross_platform_product_rows(orders, products, price_master=master)

    assert [row["reporting_order_created_date"] for row in reporting_rows] == [date(2026, 8, 7), date(2026, 8, 8), date(2026, 8, 9)]
    assert [row["reporting_sales_amount"] for row in reporting_rows] == ["20.00", "9.00", "35.70"]
    assert summarize_cross_platform_products(reporting_rows) == [
        {"seller_sku": "SKU-SHARED", "product_name": "Different source label", "unit_selling_price": Decimal("10.00"), "total_quantity": 1, "total_selling_price": Decimal("9.00"), "total_discount_given": Decimal("1.00")},
        {"seller_sku": "SKU-SHARED", "product_name": "Shared SKU first name", "unit_selling_price": Decimal("10.00"), "total_quantity": 2, "total_selling_price": Decimal("20.00"), "total_discount_given": Decimal("0.00")},
        {"seller_sku": "SKU-ZENXIN", "product_name": "ZENXIN product", "unit_selling_price": Decimal("11.90"), "total_quantity": 3, "total_selling_price": Decimal("35.70"), "total_discount_given": Decimal("0.00")},
    ]

    assert filter_cross_platform_product_rows(reporting_rows, start_date=date(2026, 8, 8), end_date=date(2026, 8, 8)) == [reporting_rows[1]]
    assert filter_cross_platform_product_rows(reporting_rows, platform="Shopee") == [reporting_rows[0]]
def test_cross_platform_dates_prefer_order_payload_then_product_source_without_hiding_missing_rows():
    orders = [
        {"platform": "Shopee", "order_id": "SHP-FALLBACK", "order_created_date": "N/A"},
        {"platform": "Lazada", "order_id": "LZD-FALLBACK", "order_date": ""},
        {"platform": "ZENXIN", "order_id": "ZNX-FALLBACK", "invoice_date": "N/A"},
        {"platform": "Shopee", "order_id": "SHP-MISSING", "order_created_date": ""},
    ]
    products = [
        {
            "platform": "Shopee",
            "order_id": "SHP-FALLBACK",
            "product_name": "Shopee fallback",
            "seller_sku": "SKU-SHP",
            "unit_price": "5.00",
            "line_subtotal": "5.00",
            "quantity": 1,
            "order_created_date": "07/08/2026 12:04",
        },
        {
            "platform": "Lazada",
            "order_id": "LZD-FALLBACK",
            "product_name": "Lazada fallback",
            "seller_sku": "SKU-LZD",
            "unit_price": "6.00",
            "paid_price": "6.00",
            "quantity": 1,
            "order_date": "08 08 2026",
        },
        {
            "platform": "ZENXIN",
            "order_id": "ZNX-FALLBACK",
            "product_name": "ZENXIN fallback",
            "seller_sku": "SKU-ZNX",
            "unit_price": "7.00",
            "line_total_inc_tax": "7.00",
            "quantity": 1,
            "invoice_date": "09/08/2026",
        },
        {
            "platform": "Shopee",
            "order_id": "SHP-MISSING",
            "product_name": "Genuinely missing date",
            "seller_sku": "SKU-MISSING-DATE",
            "unit_price": "8.00",
            "line_subtotal": "8.00",
            "quantity": 1,
        },
        {
            "platform": "Shopee",
            "order_id": "SHP-NO-SKU",
            "product_name": "No SKU stays in detail",
            "seller_sku": "N/A",
            "unit_price": "9.00",
            "line_subtotal": "9.00",
            "quantity": 1,
            "order_created_date": "10/08/2026",
        },
    ]
    reviews = [
        {
            "platform": "Shopee",
            "order_id": "SHP-REVIEW-ORDER",
            "status": "Manual Review",
            "order_payload": {
                "platform": "Shopee",
                "order_id": "SHP-REVIEW-ORDER",
                "order_created_date": "11/08/2026 09:30",
            },
            "product_payloads": [
                {
                    "platform": "Shopee",
                    "order_id": "SHP-REVIEW-ORDER",
                    "product_name": "Review order date",
                    "seller_sku": "SKU-REVIEW-ORDER",
                    "unit_price": "10.00",
                    "line_subtotal": "10.00",
                    "quantity": 1,
                }
            ],
        },
        {
            "platform": "Lazada",
            "order_id": "LZD-REVIEW-PRODUCT",
            "status": "Manual Review",
            "order_payload": {
                "platform": "Lazada",
                "order_id": "LZD-REVIEW-PRODUCT",
                "order_date": "N/A",
            },
            "product_payloads": [
                {
                    "platform": "Lazada",
                    "order_id": "LZD-REVIEW-PRODUCT",
                    "product_name": "Review product fallback",
                    "seller_sku": "SKU-REVIEW-PRODUCT",
                    "unit_price": "11.00",
                    "paid_price": "11.00",
                    "quantity": 1,
                    "order_date": "12 08 2026",
                }
            ],
        },
    ]

    detail_rows = build_cross_platform_product_rows(orders, products, reviews)
    missing_by_platform = {
        platform: sum(
            row["platform"] == platform and row["reporting_order_created_date"] is None
            for row in detail_rows
        )
        for platform in ("Shopee", "Lazada", "ZENXIN")
    }

    assert len(detail_rows) == 7
    assert missing_by_platform == {"Shopee": 1, "Lazada": 0, "ZENXIN": 0}
    assert filter_cross_platform_product_rows(detail_rows) == detail_rows
    assert len(
        filter_cross_platform_product_rows(
            detail_rows,
            start_date=date(2026, 8, 7),
            end_date=date(2026, 8, 12),
        )
    ) == 6
    assert "No SKU stays in detail" in [row["product_name"] for row in detail_rows]
    assert "N/A" not in {
        row["seller_sku"] for row in summarize_cross_platform_products(detail_rows)
    }
    assert next(
        row for row in detail_rows if row["order_id"] == "SHP-REVIEW-ORDER"
    )["reporting_order_created_date"] == date(2026, 8, 11)
    assert next(
        row for row in detail_rows if row["order_id"] == "LZD-REVIEW-PRODUCT"
    )["reporting_order_created_date"] == date(2026, 8, 12)


def test_cross_platform_summary_groups_same_price_identity_across_multiple_orders():
    master = ProductPriceMaster.from_rows([
        {"seller_sku": "SKU-SHARED", "product_name": "First source name", "unit_selling_price": "10.00"},
        {"seller_sku": "SKU-SHARED", "product_name": "Later source name", "unit_selling_price": "10.00"},
        {"seller_sku": "SKU-SHARED", "product_name": "Platform label does not split the SKU", "unit_selling_price": "10.00"},
    ])
    orders = [
        {"platform": "Shopee", "order_id": "SHP-1", "order_created_date": "07/08/2026"},
        {"platform": "Shopee", "order_id": "SHP-2", "order_created_date": "08/08/2026"},
        {"platform": "Lazada", "order_id": "LZD-1", "order_date": "09 08 2026"},
    ]
    products = [
        {"platform": "Shopee", "order_id": "SHP-1", "product_name": "First source name", "seller_sku": "SKU-SHARED", "unit_price": "10.00", "quantity": 2, "line_subtotal": "20.00", "source_line_subtotal": "20.00"},
        {"platform": "Shopee", "order_id": "SHP-2", "product_name": "Later source name", "seller_sku": "SKU-SHARED", "unit_price": "10.00", "quantity": 4, "line_subtotal": "40.00", "source_line_subtotal": "40.00"},
        {"platform": "Lazada", "order_id": "LZD-1", "product_name": "Platform label does not split the SKU", "seller_sku": "SKU-SHARED", "unit_price": "9.00", "quantity": 1, "paid_price": "9.00"},
    ]

    summary_rows = summarize_cross_platform_products(build_cross_platform_product_rows(orders, products, price_master=master))

    assert summary_rows == [
        {"seller_sku": "SKU-SHARED", "product_name": "First source name", "unit_selling_price": Decimal("10.00"), "total_quantity": 2, "total_selling_price": Decimal("20.00"), "total_discount_given": Decimal("0.00")},
        {"seller_sku": "SKU-SHARED", "product_name": "Later source name", "unit_selling_price": Decimal("10.00"), "total_quantity": 4, "total_selling_price": Decimal("40.00"), "total_discount_given": Decimal("0.00")},
        {"seller_sku": "SKU-SHARED", "product_name": "Platform label does not split the SKU", "unit_selling_price": Decimal("10.00"), "total_quantity": 1, "total_selling_price": Decimal("9.00"), "total_discount_given": Decimal("1.00")},
    ]
def test_cross_platform_filters_support_independent_inclusive_from_and_to_dates():
    rows = [
        {"platform": "Shopee", "reporting_order_created_date": date(2026, 8, 7)},
        {"platform": "Lazada", "reporting_order_created_date": date(2026, 8, 8)},
        {"platform": "ZENXIN", "reporting_order_created_date": date(2026, 8, 9)},
        {"platform": "Shopee", "reporting_order_created_date": None},
    ]

    assert filter_cross_platform_product_rows(rows, start_date=date(2026, 8, 8)) == rows[1:3]
    assert filter_cross_platform_product_rows(rows, end_date=date(2026, 8, 8)) == rows[:2]
    assert filter_cross_platform_product_rows(rows) == rows
