from datetime import datetime

from openpyxl import load_workbook

from src.invoice_app.services.all_products import ALL_PRODUCT_COLUMNS, ALL_PRODUCT_FIELD_LABELS
from src.invoice_app.services.batch_service import FIELD_LABELS, PLATFORM_ORDER_FIELDS
from src.invoice_app.services.exporter import (
    export_all_products_report,
    export_platform_report,
    export_review_report,
)


def test_platform_export_is_consistently_formatted_and_typed(tmp_path):
    destination = tmp_path / "shopee-report.xlsx"
    export_platform_report(
        destination=destination,
        platform_name="Shopee",
        summary={
            "orders": 2,
            "products": 2,
            "quantity": 3,
            "gross_sales": "1250.50",
            "total_fees": "-52.25",
            "net_amount": "1198.25",
        },
        orders=[
            {
                "platform": "Shopee",
                "order_id": "00123456789",
                "order_created_date": "07/08/2026 12:04",
                "gross_sales": "1,250.50",
                "source_pdf": "order-001.pdf",
            },
            {
                "platform": "Shopee",
                "order_id": "00123456790",
                "order_created_date": "N/A",
                "gross_sales": "-52.25",
                "source_pdf": "a-very-long-source-invoice-filename-remains-readable-and-wrapped.pdf",
            },
        ],
        products=[
            {
                "platform": "Shopee",
                "order_id": "00123456789",
                "product_name": "Organic vegetable bundle",
                "seller_sku": 9555208106364.0,
                "quantity": "3",
                "unit_price": "12.50",
            }
        ],
        order_columns=["platform", "order_id", "order_created_date", "gross_sales", "source_pdf"],
        product_columns=["platform", "order_id", "product_name", "seller_sku", "quantity", "unit_price"],
        column_labels={
            "platform": "Platform",
            "order_id": "Order ID",
            "order_created_date": "Order Created Date",
            "gross_sales": "Gross Sales",
            "source_pdf": "Source PDF",
            "product_name": "Product Name",
            "seller_sku": "Seller SKU",
            "quantity": "Quantity",
            "unit_price": "Unit Price",
        },
    )

    workbook = load_workbook(destination)
    assert workbook.sheetnames == ["Summary", "Orders", "Products"]
    assert workbook.active.title == "Summary"

    orders = workbook["Orders"]
    assert orders.freeze_panes == "A4"
    assert orders.auto_filter.ref == "A3:E5"
    assert orders.sheet_view.showGridLines is True
    assert orders.sheet_view.zoomScale == 100
    assert orders.sheet_view.zoomScaleNormal == 100
    assert orders["A1"].value == "SALES TRANSACTIONS - SHOPEE"
    assert not orders.merged_cells.ranges
    assert orders["A1"].fill.fgColor.rgb == "00FFFFFF"
    assert orders["A1"].font.name == "Calibri"
    assert orders["A1"].font.sz == 14
    assert orders["A1"].font.bold is True
    assert orders["A1"].font.color.type == "rgb"
    assert orders["A1"].font.color.rgb == "001F4E78"
    assert orders.row_dimensions[1].height == 24
    assert orders.row_dimensions[2].height == 18
    assert orders["A2"].border.left.style == "thin"
    assert orders["A2"].border.left.color.rgb == "00000000"
    assert orders["A3"].fill.fgColor.rgb == "001F4E78"
    assert orders["A3"].font.name == "Calibri"
    assert orders["A3"].font.sz == 11
    assert orders["A3"].font.bold is True
    assert orders["A3"].font.color.type == "rgb"
    assert orders["A3"].font.color.rgb == "00FFFFFF"
    assert orders["A3"].alignment.horizontal == "center"
    assert orders.row_dimensions[3].height == 20
    assert orders.column_dimensions["B"].width >= 16
    assert orders.column_dimensions["E"].width >= 24

    assert orders["B4"].value == "00123456789"
    assert orders["B4"].number_format == "@"
    assert isinstance(orders["C4"].value, datetime)
    assert orders["C4"].number_format == "dd/mm/yyyy hh:mm"
    assert orders["D4"].value == 1250.5
    assert orders["D4"].number_format == '#,##0.00;[Red]-#,##0.00'
    assert orders["D4"].alignment.horizontal == "left"
    assert orders["E5"].alignment.wrap_text is True
    assert orders["A4"].fill.fgColor.rgb == "00FFFFFF"
    assert orders["A4"].font.name == "Calibri"
    assert orders["A4"].font.sz == 11
    assert orders["A4"].border.left.style == "thin"
    assert orders["A4"].border.left.color.rgb == "00000000"

    products = workbook["Products"]
    assert products["D4"].value == "9555208106364"
    assert products["D4"].number_format == "@"
    assert products["E4"].value == 3
    assert products["E4"].number_format == "#,##0"
    assert products["F4"].value == 12.5

    summary = workbook["Summary"]
    assert summary["A4"].value == "Platform"
    assert summary["B4"].value == "Shopee"
    assert summary["B5"].value == 2
    assert summary["B8"].value == 1250.5
    assert summary["B8"].number_format == '#,##0.00;[Red]-#,##0.00'


def test_review_export_uses_same_standard_format(tmp_path):
    destination = tmp_path / "reviews.xlsx"
    export_review_report(
        [
            {
                "Batch ID": "batch-001",
                "Order ID": "000123",
                "Status": "Manual Review",
                "Reason": "PDF file could not be read.",
                "Processing Timestamp": "2026-08-20T09:30:00+00:00",
                "order_payload": {"delivery_fee": "4.90"},
                "product_payloads": [{"product_name": "Internal only"}],
            }
        ],
        destination,
    )

    workbook = load_workbook(destination)
    worksheet = workbook["Manual Review"]
    assert worksheet.freeze_panes == "A4"
    assert worksheet.auto_filter.ref == "A3:E4"
    assert worksheet["A1"].value == "MANUAL REVIEW REPORT"
    assert worksheet["A3"].fill.fgColor.rgb == "001F4E78"
    assert worksheet["B4"].value == "000123"
    assert worksheet["B4"].number_format == "@"
    assert isinstance(worksheet["E4"].value, datetime)
    assert worksheet["E4"].number_format == "dd/mm/yyyy hh:mm"
    assert worksheet["D4"].alignment.wrap_text is True
    assert "order_payload" not in {str(cell.value) for cell in worksheet[3]}
    assert "product_payloads" not in {str(cell.value) for cell in worksheet[3]}


def test_shopee_canonical_order_column_selection_exports_labels_and_types(tmp_path):
    selected_order_columns = [
        "platform",
        "order_id",
        "order_income",
        "income_type",
        "payment_status",
        "shipping_fee_rebate_from_shopee",
    ]
    for column in selected_order_columns:
        assert column in PLATFORM_ORDER_FIELDS["Shopee"]

    destination = tmp_path / "shopee-canonical-contract.xlsx"
    export_platform_report(
        destination=destination,
        platform_name="Shopee",
        summary={},
        orders=[
            {
                "platform": "Shopee",
                "order_id": "SHP-CANONICAL-1",
                "order_income": "24.53",
                "income_type": "Estimated",
                "payment_status": "Pending",
                "shipping_fee_rebate_from_shopee": "3.20",
            }
        ],
        products=[],
        order_columns=selected_order_columns,
        product_columns=["platform", "order_id"],
        column_labels=FIELD_LABELS,
    )

    workbook = load_workbook(destination)
    orders = workbook["Orders"]
    assert [cell.value for cell in orders[3]] == [FIELD_LABELS[column] for column in selected_order_columns]
    assert orders["C4"].value == 24.53
    assert orders["C4"].number_format == '#,##0.00;[Red]-#,##0.00'
    assert orders["D4"].value == "Estimated"
    assert orders["D4"].number_format == "@"
    assert orders["E4"].value == "Pending"
    assert orders["F4"].value == 3.2


def test_export_format_remains_stable_for_empty_and_large_datasets(tmp_path):
    empty_destination = tmp_path / "empty.xlsx"
    export_platform_report(
        destination=empty_destination,
        platform_name="Lazada",
        summary={},
        orders=[],
        products=[],
        order_columns=["platform", "order_id", "total"],
        product_columns=["platform", "order_id", "product_name", "quantity"],
        column_labels={
            "platform": "Platform",
            "order_id": "Order ID",
            "total": "Total",
            "product_name": "Product Name",
            "quantity": "Quantity",
        },
    )

    empty_workbook = load_workbook(empty_destination)
    empty_orders = empty_workbook["Orders"]
    assert empty_orders.max_row == 3
    assert empty_orders.auto_filter.ref == "A3:C3"
    assert empty_orders["A3"].fill.fgColor.rgb == "001F4E78"

    large_destination = tmp_path / "large.xlsx"
    large_orders = [
        {
            "platform": "Shopee",
            "order_id": f"ORDER-{index:05d}",
            "gross_sales": f"{index + 0.25:.2f}",
            "source_pdf": "large-batch.pdf",
        }
        for index in range(1500)
    ]
    export_platform_report(
        destination=large_destination,
        platform_name="Shopee",
        summary={"orders": len(large_orders)},
        orders=large_orders,
        products=[],
        order_columns=["platform", "order_id", "gross_sales", "source_pdf"],
        product_columns=["platform", "order_id", "product_name", "quantity"],
        column_labels={
            "platform": "Platform",
            "order_id": "Order ID",
            "gross_sales": "Gross Sales",
            "source_pdf": "Source PDF",
            "product_name": "Product Name",
            "quantity": "Quantity",
        },
    )

    large_workbook = load_workbook(large_destination, read_only=False)
    large_sheet = large_workbook["Orders"]
    assert large_sheet.max_row == 1503
    assert large_sheet.auto_filter.ref == "A3:D1503"
    assert large_sheet["B1503"].value == "ORDER-01499"
    assert large_sheet["C1503"].value == 1499.25
    assert large_sheet["C1503"].number_format == '#,##0.00;[Red]-#,##0.00'


def test_review_export_excludes_duplicate_and_unsupported_rows(tmp_path):
    destination = tmp_path / "actionable-reviews.xlsx"
    export_review_report(
        [
            {"Order ID": "MR-1", "Status": "Manual Review", "Reason": "Product Count Mismatch"},
            {"Order ID": "DUP-1", "Status": "Duplicate Skipped", "Reason": "Duplicate Order"},
            {"Order ID": "N/A", "Status": "Unsupported", "Reason": "Not recognized"},
        ],
        destination,
    )

    worksheet = load_workbook(destination)["Manual Review"]
    headers = [cell.value for cell in worksheet[3]]
    order_id_column = headers.index("Order ID") + 1
    assert worksheet.max_row == 4
    assert worksheet.cell(row=4, column=order_id_column).value == "MR-1"


def test_sku_columns_are_exported_as_text_without_scientific_notation(tmp_path):
    destination = tmp_path / "all-products-sku-text.xlsx"
    products = [
        {
            "product_name": "Simply Natural Raw Almonds",
            "unit_price": "8.90",
            "seller_sku": 9555208109860,
            "quantity": 27,
            "platform": "Shopee",
        },
        {
            "product_name": "Simply Natural Italy Pasta",
            "unit_price": "38.90",
            "seller_sku": "9555208109723-italy",
            "quantity": 2,
            "platform": "Shopee",
        },
    ]

    export_all_products_report(
        destination=destination,
        products=products,
        product_columns=ALL_PRODUCT_COLUMNS,
        column_labels=ALL_PRODUCT_FIELD_LABELS,
    )

    workbook = load_workbook(destination)
    worksheet = workbook["Products"]

    headers = [cell.value for cell in worksheet[3]]
    sku_col_idx = headers.index("Seller SKU #") + 1

    cell_numeric_sku = worksheet.cell(row=4, column=sku_col_idx)
    assert cell_numeric_sku.value == "9555208109860"
    assert cell_numeric_sku.data_type == "s"
    assert cell_numeric_sku.number_format == "@"

    cell_str_sku = worksheet.cell(row=5, column=sku_col_idx)
    assert cell_str_sku.value == "9555208109723-italy"
    assert cell_str_sku.data_type == "s"
    assert cell_str_sku.number_format == "@"
