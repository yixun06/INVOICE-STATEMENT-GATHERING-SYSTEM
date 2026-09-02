from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from enum import Enum
from functools import lru_cache
from pathlib import Path
import re
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from src.invoice_app.utils.normalize import (
    normalize_match_text,
    normalize_sku_text,
    normalize_whitespace,
)


class PriceLookupStatus(str, Enum):
    MATCHED_BY_SKU = "matched_by_sku"
    MATCHED_BY_SKU_NAME_VARIATION = "matched_by_sku_name_variation"
    MATCHED_BY_PARENT_SKU = "matched_by_parent_sku"
    PRICE_CONFIRMED_IDENTITY_AMBIGUOUS = "price_confirmed_identity_ambiguous"
    MATCHED_BY_PARENT_SKU_NAME_VARIATION = "matched_by_parent_sku_name_variation"
    MATCHED = "matched"
    MATCHED_BY_ALIAS = "matched_by_alias"
    MATCHED_BY_NAME_VARIATION = "matched_by_name_variation"
    PRICING_CONFLICT = "pricing_conflict"
    PRICE_NOT_FOUND = "price_not_found"


@dataclass(frozen=True)
class ProductPriceMasterRecord:
    seller_sku: str
    parent_sku: str
    product_name: str
    variation_name: str
    unit_selling_price: Decimal
    source_row: int


@dataclass(frozen=True)
class ProductPriceMasterMetadata:
    source_filename: str
    sheet_name: str
    header_row: int | None
    workbook_row_count: int
    candidate_row_count: int
    loaded_row_count: int
    invalid_price_row_count: int


@dataclass(frozen=True)
class ProductPriceLookupResult:
    status: PriceLookupStatus
    unit_selling_price: Decimal | None
    matched_by: str | None
    matched_sku: str | None
    matched_parent_sku: str | None
    matched_product_name: str | None
    matched_variation_name: str | None
    source_metadata: ProductPriceMasterMetadata
    source_rows: tuple[int, ...]
    reason: str | None = None

    alias_rule: str | None = None

_REQUIRED_HEADERS = {
    "product_name": "product name",
    "variation_name": "variation name",
    "parent_sku": "parent sku",
    "seller_sku": "sku",
    "unit_selling_price": "price",
}
_STRICT_PRICE = re.compile(
    r"^(?:RM\s*)?(-?(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d+)?)$",
    re.IGNORECASE,
)


class ProductPriceMaster:
    """Read-only Shopee Product Listing lookup index."""

    def __init__(
        self,
        records: Iterable[ProductPriceMasterRecord],
        metadata: ProductPriceMasterMetadata,
    ) -> None:
        self.metadata = metadata
        self.records = tuple(records)
        self._seller_sku_index = _build_index(self.records, "seller_sku")
        self._parent_sku_index = _build_index(self.records, "parent_sku")

    @classmethod
    def from_xlsx(cls, source_path: str | Path) -> ProductPriceMaster:
        """Load the Product Listing on demand into an in-memory core-service index."""
        path = Path(source_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            header_row, columns = _find_listing_headers(worksheet)
            records: list[ProductPriceMasterRecord] = []
            candidate_rows = 0
            invalid_prices = 0

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                values = {
                    field: row[index] if index < len(row) else None
                    for field, index in columns.items()
                }
                seller_sku = normalize_sku_text(values["seller_sku"])
                parent_sku = normalize_sku_text(values["parent_sku"])
                if not seller_sku and not parent_sku:
                    continue

                candidate_rows += 1
                price = _strict_decimal(values["unit_selling_price"])
                if price is None:
                    invalid_prices += 1
                    continue

                records.append(
                    ProductPriceMasterRecord(
                        seller_sku=seller_sku,
                        parent_sku=parent_sku,
                        product_name=_display_text(values["product_name"]),
                        variation_name=_display_text(values["variation_name"]),
                        unit_selling_price=price,
                        source_row=row_number,
                    )
                )

            metadata = ProductPriceMasterMetadata(
                source_filename=path.name,
                sheet_name=worksheet.title,
                header_row=header_row,
                workbook_row_count=worksheet.max_row,
                candidate_row_count=candidate_rows,
                loaded_row_count=len(records),
                invalid_price_row_count=invalid_prices,
            )
            return cls(records, metadata)
        finally:
            workbook.close()

    @classmethod
    def from_rows(
        cls,
        rows: Iterable[Mapping[str, Any]],
        *,
        source_filename: str = "<in-memory>",
    ) -> ProductPriceMaster:
        """Build a synthetic master without reading an XLSX."""
        records: list[ProductPriceMasterRecord] = []
        candidate_rows = 0
        invalid_prices = 0
        for source_row, row in enumerate(rows, start=1):
            seller_sku = normalize_sku_text(row.get("seller_sku"))
            parent_sku = normalize_sku_text(row.get("parent_sku"))
            if not seller_sku and not parent_sku:
                continue
            candidate_rows += 1
            price = _strict_decimal(row.get("unit_selling_price"))
            if price is None:
                invalid_prices += 1
                continue
            records.append(
                ProductPriceMasterRecord(
                    seller_sku=seller_sku,
                    parent_sku=parent_sku,
                    product_name=_display_text(row.get("product_name")),
                    variation_name=_display_text(row.get("variation_name")),
                    unit_selling_price=price,
                    source_row=source_row,
                )
            )
        return cls(
            records,
            ProductPriceMasterMetadata(
                source_filename=source_filename,
                sheet_name="<in-memory>",
                header_row=None,
                workbook_row_count=candidate_rows,
                candidate_row_count=candidate_rows,
                loaded_row_count=len(records),
                invalid_price_row_count=invalid_prices,
            ),
        )

    def lookup(
        self,
        *,
        seller_sku: str | None,
        parent_sku: str | None = None,
        product_name: str | None = None,
        variation_name: str | None = None,
    ) -> ProductPriceLookupResult:
        """Resolve one price only when it identifies one master product identity."""
        requested_seller_sku = normalize_sku_text(seller_sku)
        if _is_invalid_sku(requested_seller_sku):
            return self._resolve_name_variation(product_name, variation_name)

        candidates = self._sku_candidates(requested_seller_sku)
        if candidates:
            return self._resolve_sku_candidates(
                candidates,
                product_name=product_name,
                variation_name=variation_name,
            )

        if requested_seller_sku.endswith("-Less"):
            alias_candidates = self._sku_candidates(
                requested_seller_sku.removesuffix("-Less")
            )
            if alias_candidates:
                return self._resolve_sku_candidates(
                    alias_candidates,
                    product_name=product_name,
                    variation_name=variation_name,
                    alias_rule="REMOVE_SUFFIX_LESS",
                )
        return self._not_found(
            f"No Product Listing SKU or Parent SKU matches '{requested_seller_sku}'."
        )

    def _sku_candidates(
        self,
        requested_sku: str,
    ) -> tuple[ProductPriceMasterRecord, ...]:
        return _unique_records(
            self._seller_sku_index.get(requested_sku, ())
            + self._parent_sku_index.get(requested_sku, ())
        )

    def _resolve_sku_candidates(
        self,
        candidates: tuple[ProductPriceMasterRecord, ...],
        *,
        product_name: str | None,
        variation_name: str | None,
        alias_rule: str | None = None,
    ) -> ProductPriceLookupResult:
        if _has_one_identity(candidates):
            return self._matched_identity(
                candidates,
                status=(
                    PriceLookupStatus.MATCHED_BY_ALIAS
                    if alias_rule
                    else PriceLookupStatus.MATCHED
                ),
                alias_rule=alias_rule,
            )

        narrowed = _match_name_variation(candidates, product_name, variation_name)
        if narrowed and _has_one_identity(narrowed):
            return self._matched_identity(
                narrowed,
                status=(
                    PriceLookupStatus.MATCHED_BY_ALIAS
                    if alias_rule
                    else PriceLookupStatus.MATCHED
                ),
                alias_rule=alias_rule,
            )
        variation_narrowed = _match_variation(candidates, variation_name)
        if variation_narrowed and _has_one_identity(variation_narrowed):
            return self._matched_identity(
                variation_narrowed,
                status=(
                    PriceLookupStatus.MATCHED_BY_ALIAS
                    if alias_rule
                    else PriceLookupStatus.MATCHED
                ),
                alias_rule=alias_rule,
                matched_by="sku_unique_exact_variation",
            )

        if _has_one_price(candidates):
            return self._price_confirmed_identity_ambiguous(
                candidates,
                alias_rule=alias_rule,
            )


        if narrowed:
            evidence = narrowed
            reason = (
                "Multiple master identities remain after exact Product Name / "
                "Variation matching."
            )
        elif _has_disambiguation_input(product_name, variation_name):
            evidence = candidates
            reason = (
                "Product Name / Variation did not identify one candidate master identity."
            )
        else:
            evidence = candidates
            reason = (
                "Multiple master identities require exact Product Name / Variation "
                "disambiguation."
            )
        return self._conflict(evidence, reason)

    def _resolve_name_variation(
        self,
        product_name: str | None,
        variation_name: str | None,
    ) -> ProductPriceLookupResult:
        if not _match_text(product_name):
            return self._not_found(
                "Seller SKU is blank, N/A, or exp and Product Name is unavailable."
            )
        candidates = _match_name_variation(self.records, product_name, variation_name)
        if not candidates:
            return self._not_found(
                "No Product Listing identity exactly matches Product Name / Variation."
            )
        if _has_one_identity(candidates):
            return self._matched_identity(
                candidates,
                status=PriceLookupStatus.MATCHED_BY_NAME_VARIATION,
            )
        return self._conflict(
            candidates,
            "Product Name / Variation matches multiple master identities.",
        )

    def _matched_identity(
        self,
        records: tuple[ProductPriceMasterRecord, ...],
        *,
        status: PriceLookupStatus,
        alias_rule: str | None = None,
        matched_by: str | None = None,
    ) -> ProductPriceLookupResult:
        record = min(records, key=lambda item: item.source_row)
        return ProductPriceLookupResult(
            status=status,
            unit_selling_price=record.unit_selling_price,
            matched_by=matched_by or status.value,
            matched_sku=record.seller_sku or None,
            matched_parent_sku=record.parent_sku or None,
            matched_product_name=record.product_name or None,
            matched_variation_name=record.variation_name or None,
            source_metadata=self.metadata,
            source_rows=tuple(sorted(item.source_row for item in records)),
            alias_rule=alias_rule,
        )


    def _price_confirmed_identity_ambiguous(
        self,
        records: tuple[ProductPriceMasterRecord, ...],
        *,
        alias_rule: str | None = None,
    ) -> ProductPriceLookupResult:
        price = records[0].unit_selling_price
        return ProductPriceLookupResult(
            status=PriceLookupStatus.PRICE_CONFIRMED_IDENTITY_AMBIGUOUS,
            unit_selling_price=price,
            matched_by="unique_candidate_price",
            matched_sku=None,
            matched_parent_sku=None,
            matched_product_name=None,
            matched_variation_name=None,
            source_metadata=self.metadata,
            source_rows=tuple(sorted(item.source_row for item in records)),
            alias_rule=alias_rule,
            reason=(
                "Candidate identities remain unresolved, but all exact SKU / Parent SKU "
                "candidates share one unique Unit Selling Price."
            ),
        )

    def _matched(
        self,
        records: tuple[ProductPriceMasterRecord, ...],
        match_scope: str,
        *,
        disambiguated: bool,
    ) -> ProductPriceLookupResult:
        record = min(records, key=lambda item: item.source_row)
        status = _matched_status(match_scope, disambiguated)
        return ProductPriceLookupResult(
            status=status,
            unit_selling_price=record.unit_selling_price,
            matched_by=status.value,
            matched_sku=record.seller_sku or None,
            matched_parent_sku=record.parent_sku or None,
            matched_product_name=record.product_name or None,
            matched_variation_name=record.variation_name or None,
            source_metadata=self.metadata,
            source_rows=tuple(sorted(item.source_row for item in records)),
        )

    def _conflict(
        self,
        records: tuple[ProductPriceMasterRecord, ...],
        reason: str,
    ) -> ProductPriceLookupResult:
        return ProductPriceLookupResult(
            status=PriceLookupStatus.PRICING_CONFLICT,
            unit_selling_price=None,
            matched_by=None,
            matched_sku=None,
            matched_parent_sku=None,
            matched_product_name=None,
            matched_variation_name=None,
            source_metadata=self.metadata,
            source_rows=tuple(sorted(item.source_row for item in records)),
            reason=reason,
        )

    def _not_found(self, reason: str) -> ProductPriceLookupResult:
        return ProductPriceLookupResult(
            status=PriceLookupStatus.PRICE_NOT_FOUND,
            unit_selling_price=None,
            matched_by=None,
            matched_sku=None,
            matched_parent_sku=None,
            matched_product_name=None,
            matched_variation_name=None,
            source_metadata=self.metadata,
            source_rows=(),
            reason=reason,
        )


def load_shopee_product_price_master(source_path: str | Path) -> ProductPriceMaster:
    """Cache the source by path, so normal callers do not reread it per PDF."""
    return _load_cached_price_master(str(Path(source_path).resolve()))


@lru_cache(maxsize=4)
def _load_cached_price_master(resolved_path: str) -> ProductPriceMaster:
    return ProductPriceMaster.from_xlsx(resolved_path)


def _build_index(
    records: Iterable[ProductPriceMasterRecord],
    attribute: str,
) -> dict[str, tuple[ProductPriceMasterRecord, ...]]:
    index: dict[str, list[ProductPriceMasterRecord]] = defaultdict(list)
    for record in records:
        value = getattr(record, attribute)
        if value:
            index[value].append(record)
    return {key: tuple(value) for key, value in index.items()}


def _find_listing_headers(worksheet: Any) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        headers = {
            _header_text(value): index
            for index, value in enumerate(row)
            if _header_text(value)
        }
        if all(expected in headers for expected in _REQUIRED_HEADERS.values()):
            return row_number, {
                field: headers[expected] for field, expected in _REQUIRED_HEADERS.items()
            }
        if row_number >= 50:
            break
    raise ValueError("Shopee Product Listing required headers were not found.")


def _header_text(value: Any) -> str:
    return normalize_whitespace(str(value or "")).casefold()


def _sku_text(value: Any) -> str:
    """Keep identifiers as text: no numeric parsing or case coercion."""
    return "" if value is None else str(value).strip()


def _display_text(value: Any) -> str:
    return normalize_whitespace(str(value or ""))


def _strict_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    match = _STRICT_PRICE.fullmatch(str(value).strip())
    if not match:
        return None
    try:
        return Decimal(match.group(1).replace(",", ""))
    except InvalidOperation:
        return None


def _has_one_price(records: tuple[ProductPriceMasterRecord, ...]) -> bool:
    return len({record.unit_selling_price for record in records}) == 1


def _has_disambiguation_input(
    product_name: str | None,
    variation_name: str | None,
) -> bool:
    return bool(_match_text(product_name) or _match_text(variation_name))


def _match_name_variation(
    records: tuple[ProductPriceMasterRecord, ...],
    product_name: str | None,
    variation_name: str | None,
) -> tuple[ProductPriceMasterRecord, ...]:
    product_key = _match_text(product_name)
    variation_key = _match_text(variation_name)
    if not product_key and not variation_key:
        return ()
    return tuple(
        record
        for record in records
        if (not product_key or _match_text(record.product_name) == product_key)
        and (not variation_key or _match_text(record.variation_name) == variation_key)
    )


def _match_variation(
    records: tuple[ProductPriceMasterRecord, ...],
    variation_name: str | None,
) -> tuple[ProductPriceMasterRecord, ...]:
    variation_key = _match_text(variation_name)
    if not variation_key:
        return ()
    return tuple(
        record for record in records if _match_text(record.variation_name) == variation_key
    )
def _match_text(value: str | None) -> str:
    return normalize_match_text(value)


def _is_invalid_sku(value: str) -> bool:
    return not value or value.casefold() in {"n/a", "exp"}


def _has_one_identity(records: tuple[ProductPriceMasterRecord, ...]) -> bool:
    return len({_identity_key(record) for record in records}) == 1


def _identity_key(record: ProductPriceMasterRecord) -> tuple[str, str, str, str, Decimal]:
    return (
        normalize_sku_text(record.seller_sku),
        normalize_sku_text(record.parent_sku),
        _match_text(record.product_name),
        _match_text(record.variation_name),
        record.unit_selling_price,
    )



def _unique_records(records: Iterable[ProductPriceMasterRecord]) -> tuple[ProductPriceMasterRecord, ...]:
    return tuple(dict.fromkeys(records))


def _matched_status(match_scope: str, disambiguated: bool) -> PriceLookupStatus:
    if match_scope in {"seller_sku", "seller_sku_or_parent_sku"}:
        return (
            PriceLookupStatus.MATCHED_BY_SKU_NAME_VARIATION
            if disambiguated
            else PriceLookupStatus.MATCHED_BY_SKU
        )
    return (
        PriceLookupStatus.MATCHED_BY_PARENT_SKU_NAME_VARIATION
        if disambiguated
        else PriceLookupStatus.MATCHED_BY_PARENT_SKU
    )
