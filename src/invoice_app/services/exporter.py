from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from math import ceil
from pathlib import Path
import re
from typing import Any, Sequence

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
TITLE_FILL = WHITE_FILL
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11, color="000000")
GRID_SIDE = Side(style="thin", color="000000")
GRID_BORDER = Border(left=GRID_SIDE, right=GRID_SIDE, top=GRID_SIDE, bottom=GRID_SIDE)
MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00'
INTEGER_FORMAT = "#,##0"
DATE_FORMAT = "dd/mm/yyyy"
DATETIME_FORMAT = "dd/mm/yyyy hh:mm"

_TEXT_HEADERS = {
    "batch id",
    "order id",
    "invoice number",
    "seller sku",
    "shop sku",
    "voucher code",
    "source pdf",
}
_INTEGER_HEADERS = {"quantity", "qty", "orders", "products", "pdf count", "product rows", "total quantity"}
_MONEY_WORDS = {
    "amount",
    "discount",
    "fee",
    "fees",
    "income",
    "paid",
    "price",
    "released",
    "sales",
    "subtotal",
    "total",
    "voucher",
}
_NON_MONEY_HEADERS = {"income type", "payment method", "voucher code", "voucher funded by", "voucher type"}
_LONG_TEXT_HEADERS = {"product name", "reason", "source pdf"}
_NUMERIC_PATTERN = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)$")
_DATE_PATTERNS = (
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%d %b %Y",
    "%d %B %Y",
)


def export_platform_report(
    destination: str | Path,
    platform_name: str,
    summary: dict[str, Any],
    orders: Sequence[dict[str, Any]],
    products: Sequence[dict[str, Any]],
    order_columns: Sequence[str] | None = None,
    product_columns: Sequence[str] | None = None,
    column_labels: dict[str, str] | None = None,
) -> Path:
    filepath = Path(destination)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    summary_df = _summary_frame(platform_name, summary)
    orders_df = _frame_for_export(orders, order_columns, column_labels)
    products_df = _frame_for_export(products, product_columns, column_labels)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        orders_df.to_excel(writer, index=False, sheet_name="Orders")
        products_df.to_excel(writer, index=False, sheet_name="Products")
        format_excel_workbook(writer.book, platform_name=platform_name)
        _format_summary_values(writer.book["Summary"])

    return filepath


def export_all_products_report(
    destination: str | Path,
    products: Sequence[dict[str, Any]],
    product_columns: Sequence[str],
    column_labels: dict[str, str],
) -> Path:
    """Export the All view as one formatted Products worksheet only."""
    filepath = Path(destination)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    products_df = _frame_for_export(products, product_columns, column_labels)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        products_df.to_excel(writer, index=False, sheet_name="Products", startrow=2)
        worksheet = writer.book["Products"]
        _format_all_products_worksheet(worksheet)

    return filepath


def export_review_report(reviews: Sequence[dict[str, Any]], destination: str | Path) -> Path:
    actionable_reviews = [review for review in reviews if _is_manual_review_export_row(review)]
    dataframe = pd.DataFrame(actionable_reviews)
    internal_columns = [
        column
        for column in dataframe.columns
        if str(column).replace("_", "").replace(" ", "").casefold()
        in {"orderpayload", "productpayloads"}
    ]
    dataframe = dataframe.drop(columns=internal_columns, errors="ignore")
    filepath = Path(destination)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Manual Review")
        format_excel_workbook(writer.book)
    return filepath


def _is_manual_review_export_row(review: dict[str, Any]) -> bool:
    status = str(review.get("status", review.get("Status", ""))).strip().casefold()
    reason = str(review.get("reason", review.get("Reason", ""))).strip().casefold()
    return status in {"", "manual review"} and reason not in {
        "duplicate order",
        "duplicate order id in current batch for the same platform.",
    }


def format_excel_workbook(workbook: Workbook, platform_name: str | None = None) -> None:
    """Apply a compact, filterable report layout to every exported sheet."""
    workbook.properties.creator = "InvoiceGather"
    workbook.properties.title = "InvoiceGather export"
    for worksheet in workbook.worksheets:
        worksheet.insert_rows(1, amount=2)
        worksheet.sheet_view.showGridLines = True
        worksheet.sheet_view.zoomScale = 100
        worksheet.sheet_view.zoomScaleNormal = 100
        worksheet.freeze_panes = "A4"
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_title_rows = "1:3"
        _format_worksheet(worksheet, platform_name)

    if "Summary" in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index("Summary")


def _summary_frame(platform_name: str, summary: dict[str, Any]) -> pd.DataFrame:
    rows = [{"Metric": "Platform", "Value": platform_name}]
    rows.extend({"Metric": _display_label(key), "Value": value} for key, value in summary.items())
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _display_label(value: str) -> str:
    label = value.replace("_", " ").strip().title()
    return label.replace("Pdf", "PDF").replace("Sku", "SKU")


def _frame_for_export(
    rows: Sequence[dict[str, Any]],
    columns: Sequence[str] | None,
    column_labels: dict[str, str] | None = None,
) -> pd.DataFrame:
    dataframe = pd.DataFrame(list(rows))
    if not columns:
        return dataframe.rename(columns=column_labels or {})
    for column in columns:
        if column not in dataframe.columns:
            dataframe[column] = ""
    return dataframe[list(columns)].rename(columns=column_labels or {})


def _format_worksheet(worksheet: Worksheet, platform_name: str | None) -> None:
    if worksheet.max_column < 1:
        return

    header_row = 3
    for row_index in (1, 2):
        for cell in worksheet[row_index]:
            cell.fill = WHITE_FILL
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = GRID_BORDER

    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = _report_title(worksheet.title, platform_name)
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = GRID_BORDER
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 18

    headers: list[str] = []
    for cell in worksheet[header_row]:
        header = str(cell.value or "").strip()
        headers.append(header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER

    worksheet.row_dimensions[header_row].height = 20
    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(worksheet.max_column)}{max(worksheet.max_row, header_row)}"
    )

    column_widths = {
        column_index: _column_width(worksheet, column_index, header)
        for column_index, header in enumerate(headers, start=1)
    }
    for column_index, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            kind = _column_kind(header)
            _standardize_cell(cell, kind)
            cell.fill = WHITE_FILL
            cell.font = BODY_FONT
            cell.border = GRID_BORDER
            cell.alignment = _body_alignment(header, kind)
        worksheet.row_dimensions[row_index].height = _row_height(
            worksheet,
            row_index,
            headers,
            column_widths,
        )


def _format_all_products_worksheet(worksheet: Worksheet) -> None:
    """Apply the compact product-list format without platform report sheets."""
    worksheet.sheet_view.showGridLines = True
    worksheet.sheet_view.zoomScale = 100
    worksheet.sheet_view.zoomScaleNormal = 100
    worksheet.freeze_panes = "A4"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = "1:3"

    header_row = 3
    for row_index in (1, 2):
        for cell in worksheet[row_index]:
            cell.fill = WHITE_FILL
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = GRID_BORDER

    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = "PRODUCT DETAILS - ALL"
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = GRID_BORDER
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 18

    headers: list[str] = []
    for cell in worksheet[header_row]:
        header = str(cell.value or "").strip()
        headers.append(header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER

    worksheet.row_dimensions[header_row].height = 20
    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(worksheet.max_column)}{max(worksheet.max_row, header_row)}"
    )
    column_widths = {
        column_index: _column_width(worksheet, column_index, header)
        for column_index, header in enumerate(headers, start=1)
    }
    for column_index, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            kind = _column_kind(header)
            _standardize_cell(cell, kind)
            cell.fill = WHITE_FILL
            cell.font = BODY_FONT
            cell.border = GRID_BORDER
            cell.alignment = _body_alignment(header, kind)
        worksheet.row_dimensions[row_index].height = _row_height(
            worksheet,
            row_index,
            headers,
            column_widths,
        )


def _format_summary_values(worksheet: Worksheet) -> None:
    for row_index in range(4, worksheet.max_row + 1):
        metric = str(worksheet.cell(row=row_index, column=1).value or "")
        value_cell = worksheet.cell(row=row_index, column=2)
        kind = _column_kind(metric)
        _standardize_cell(value_cell, kind)
        value_cell.alignment = _body_alignment(metric, kind)


def _report_title(sheet_name: str, platform_name: str | None) -> str:
    platform = str(platform_name or "").strip().upper()
    if sheet_name == "Orders":
        prefix = "SALES TRANSACTIONS"
    elif sheet_name == "Products":
        prefix = "PRODUCT DETAILS"
    elif sheet_name == "Manual Review":
        prefix = "MANUAL REVIEW REPORT"
    else:
        prefix = "EXPORT SUMMARY"
    return f"{prefix} - {platform}" if platform else prefix


def _column_kind(header: str) -> str:
    normalized = header.strip().lower()
    words = set(re.findall(r"[a-z]+", normalized))
    if normalized in _TEXT_HEADERS or normalized.endswith(" id") or normalized.endswith(" sku"):
        return "text"
    if "date" in words or "timestamp" in words:
        return "date"
    if normalized in _INTEGER_HEADERS or normalized.endswith(" count"):
        return "integer"
    if normalized not in _NON_MONEY_HEADERS and words.intersection(_MONEY_WORDS):
        return "money"
    return "text"


def _standardize_cell(cell: Any, kind: str) -> None:
    if cell.value is None or str(cell.value).strip() == "":
        return

    if kind == "text":
        cell.value = str(cell.value)
        cell.number_format = "@"
        return

    if kind == "date":
        parsed_date = _coerce_date(cell.value)
        if parsed_date is not None:
            cell.value = parsed_date
            cell.number_format = DATETIME_FORMAT if parsed_date.time() != datetime.min.time() else DATE_FORMAT
        return

    if kind == "integer":
        parsed_integer = _coerce_integer(cell.value)
        if parsed_integer is not None:
            cell.value = parsed_integer
            cell.number_format = INTEGER_FORMAT
        return

    parsed_money = _coerce_money(cell.value)
    if parsed_money is not None:
        cell.value = parsed_money
        cell.number_format = MONEY_FORMAT


def _coerce_date(value: Any) -> datetime | None:
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().replace(tzinfo=None)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return None
    for pattern in _DATE_PATTERNS:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def _coerce_integer(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except InvalidOperation:
        return None
    if number != number.to_integral_value():
        return None
    return int(number)


def _coerce_money(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)

    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return None
    negative = text.startswith("(") and text.endswith(")")
    normalized = text.strip("()").replace(",", "").replace("RM", "").replace("rm", "").strip()
    if not _NUMERIC_PATTERN.fullmatch(normalized):
        return None
    try:
        amount = Decimal(normalized)
    except InvalidOperation:
        return None
    return float(-amount if negative else amount)


def _body_alignment(header: str, kind: str) -> Alignment:
    return Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )


def _column_width(worksheet: Worksheet, column_index: int, header: str) -> float:
    kind = _column_kind(header)
    normalized = header.strip().lower()
    measured = len(header) + 3
    for row_index in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is not None:
            measured = max(measured, len(str(value)) + 2)

    if normalized in _LONG_TEXT_HEADERS:
        return float(min(max(measured, 24), 45))
    if normalized in _TEXT_HEADERS or normalized.endswith(" id") or normalized.endswith(" sku"):
        return float(min(max(measured, 16), 28))
    if kind == "date":
        return float(min(max(measured, 16), 22))
    if kind == "money":
        return float(min(max(measured, 14), 24))
    if kind == "integer":
        return float(min(max(measured, 11), 16))
    return float(min(max(measured, 12), 30))


def _row_height(
    worksheet: Worksheet,
    row_index: int,
    headers: Sequence[str],
    column_widths: dict[int, float],
) -> float:
    required_lines = 1
    for column_index, _header in enumerate(headers, start=1):
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is None:
            continue
        usable_width = max(column_widths[column_index] - 2, 1)
        line_count = sum(max(1, ceil(len(line) / usable_width)) for line in str(value).splitlines() or [""])
        required_lines = max(required_lines, line_count)
    return float(min(max(20, required_lines * 15), 60))
