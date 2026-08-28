from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
from pathlib import Path
import posixpath
import re
from typing import Any, BinaryIO, Mapping
import warnings
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


REQUIRED_SHEETS = (
    "Summary", "Income", "Service Fee Details",
    "Shipping Fee Discrepancy", "Adjustment",
)
INCOME_COMPONENT_COLUMNS = (
    "Product Price", "Refund Amount",
    "Shipping Fee Paid by Buyer (excl. SST)",
    "Shipping Fee Charged by Logistic Provider",
    "Seller Paid Shipping Fee SST", "Shipping Rebate From Shopee",
    "Reverse Shipping Fee", "Reverse Shipping Fee SST",
    "Saver Programme Shipping Fee Savings", "Return to Seller Fee",
    "Rebate Provided by Shopee", "Voucher Sponsored by Seller",
    "Cofund Voucher Sponsored by Seller", "Coin Cashback Sponsored by Seller",
    "Cofund Coin Cashback Sponsored by Seller", "Commission Fee (incl. SST)",
    "Service Fee (Incl. SST)", "Transaction Fee (Incl. SST)",
    "AMS Commission Fee", "Saver Programme Fee (Incl. SST)",
    "Ads Escrow Top Up Fee",
)
INCOME_REQUIRED_COLUMNS = (
    "Sequence No.", "View By", "Order ID", "Product ID", "Product Name",
    "Order Creation Date", "Payout Completed Date", "Release Channel",
    "Order Type", "Total Released Amount (RM)", *INCOME_COMPONENT_COLUMNS,
)
SERVICE_FEE_REQUIRED_COLUMNS = ("Sequence No.", "Order ID")
SHIPPING_REQUIRED_COLUMNS = (
    "Order ID", "Expected Shipping Fee:",
    "Actual Shipping Fee Charged by Logistic Provider:", "Discrepancy reason",
)
ADJUSTMENT_REQUIRED_COLUMNS = (
    "Sequence No.", "Adjustment Complete Date", "Adjustment Type | Description",
    "Adjustment Reason", "Adjustment Amount", "Linked Order No.",
    "Payout Completed Date",
)


@dataclass(frozen=True)
class SourceValueIssue:
    code: str
    sheet: str
    row_number: int
    column: str
    message: str


@dataclass(frozen=True)
class SettlementIncomeRow:
    sequence_no: str
    view_by: str
    order_id: str
    product_id: str
    product_name: str
    order_creation_date: date | None
    payout_completed_date: date | None
    release_channel: str
    order_type: str
    total_released_amount: Decimal | None
    financial_components: Mapping[str, Decimal | None]
    source_values: Mapping[str, Any]
    source_row_number: int


@dataclass(frozen=True)
class ServiceFeeDetail:
    sequence_no: str
    order_id: str
    components: Mapping[str, Decimal | None]
    source_row_number: int


@dataclass(frozen=True)
class ShippingFeeDiscrepancy:
    order_id: str
    expected_shipping_fee: Decimal | None
    actual_shipping_fee: Decimal | None
    reason: str
    source_values: Mapping[str, Any]
    source_row_number: int


@dataclass(frozen=True)
class SettlementAdjustment:
    sequence_no: str
    adjustment_complete_date: date | None
    adjustment_type: str
    adjustment_reason: str
    adjustment_amount: Decimal | None
    linked_order_id: str
    payout_completed_date: date | None
    source_row_number: int


@dataclass(frozen=True)
class ParsedShopeeWeeklyStatement:
    source_filename: str
    file_hash: str
    statement_period_from: date
    statement_period_to: date
    summary_total_released: Decimal
    adjustment_control_total: Decimal
    adjustment_footer_total: Decimal | None
    income_rows: tuple[SettlementIncomeRow, ...]
    service_fee_details: tuple[ServiceFeeDetail, ...]
    shipping_fee_discrepancies: tuple[ShippingFeeDiscrepancy, ...]
    adjustments: tuple[SettlementAdjustment, ...]
    source_value_issues: tuple[SourceValueIssue, ...]
    dimension_fallback_sheets: tuple[str, ...]

    @property
    def order_rows(self) -> tuple[SettlementIncomeRow, ...]:
        return tuple(row for row in self.income_rows if row.view_by == "Order")

    @property
    def sku_rows(self) -> tuple[SettlementIncomeRow, ...]:
        return tuple(row for row in self.income_rows if row.view_by == "Sku")


class WeeklyStatementParseError(ValueError):
    def __init__(self, message: str, *, source_filename: str = "", file_hash: str = "") -> None:
        super().__init__(message)
        self.source_filename = source_filename
        self.file_hash = file_hash

def parse_shopee_weekly_statement(
    source: str | Path | bytes | bytearray | BinaryIO,
    *,
    source_filename: str | None = None,
) -> ParsedShopeeWeeklyStatement:
    workbook_bytes, filename = _read_source(source, source_filename)
    file_hash = sha256(workbook_bytes).hexdigest()
    if filename and Path(filename).suffix.casefold() != ".xlsx":
        raise WeeklyStatementParseError(
            "Shopee Weekly Statement must be an .xlsx workbook.",
            source_filename=filename, file_hash=file_hash,
        )
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore", message="Workbook contains no default style.*", category=UserWarning
            )
            workbook = load_workbook(
                BytesIO(workbook_bytes), read_only=True, data_only=True, keep_links=False
            )
    except (BadZipFile, OSError, ValueError, KeyError) as exc:
        raise WeeklyStatementParseError(
            f"Unreadable or corrupt Weekly Statement workbook: {exc}",
            source_filename=filename, file_hash=file_hash,
        ) from exc

    try:
        missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
        if missing:
            raise WeeklyStatementParseError(
                "Missing required sheet(s): " + ", ".join(missing),
                source_filename=filename, file_hash=file_hash,
            )
        fallback_sheets: list[str] = []
        summary_rows = _read_worksheet_rows(
            workbook_bytes, workbook["Summary"],
            ("3. Total Released Amount",), fallback_sheets,
        )
        income_rows = _read_worksheet_rows(
            workbook_bytes, workbook["Income"], INCOME_REQUIRED_COLUMNS, fallback_sheets,
        )
        service_rows = _read_worksheet_rows(
            workbook_bytes, workbook["Service Fee Details"],
            SERVICE_FEE_REQUIRED_COLUMNS, fallback_sheets,
        )
        shipping_rows = _read_worksheet_rows(
            workbook_bytes, workbook["Shipping Fee Discrepancy"],
            SHIPPING_REQUIRED_COLUMNS, fallback_sheets,
        )
        adjustment_rows = _read_worksheet_rows(
            workbook_bytes, workbook["Adjustment"],
            ADJUSTMENT_REQUIRED_COLUMNS, fallback_sheets,
        )
        period_from = _required_label_date(summary_rows, "From", "Summary")
        period_to = _required_label_date(summary_rows, "to", "Summary")
        summary_total = _required_label_decimal(
            summary_rows, "3. Total Released Amount", "Summary"
        )
        adjustment_control = _required_label_decimal(
            adjustment_rows, "Total Adjustment Amount", "Adjustment"
        )
        adjustment_footer = _optional_label_decimal(adjustment_rows, "Total Amount")
        issues: list[SourceValueIssue] = []
        parsed_income = _parse_income_rows(income_rows, issues)
        parsed_service = _parse_service_fee_rows(service_rows, issues)
        parsed_shipping = _parse_shipping_rows(shipping_rows, issues)
        parsed_adjustments = _parse_adjustment_rows(adjustment_rows, issues)
    except WeeklyStatementParseError as exc:
        if not exc.source_filename:
            exc.source_filename = filename
        if not exc.file_hash:
            exc.file_hash = file_hash
        raise
    finally:
        workbook.close()

    return ParsedShopeeWeeklyStatement(
        source_filename=filename,
        file_hash=file_hash,
        statement_period_from=period_from,
        statement_period_to=period_to,
        summary_total_released=summary_total,
        adjustment_control_total=adjustment_control,
        adjustment_footer_total=adjustment_footer,
        income_rows=tuple(parsed_income),
        service_fee_details=tuple(parsed_service),
        shipping_fee_discrepancies=tuple(parsed_shipping),
        adjustments=tuple(parsed_adjustments),
        source_value_issues=tuple(issues),
        dimension_fallback_sheets=tuple(fallback_sheets),
    )


def _read_source(
    source: str | Path | bytes | bytearray | BinaryIO,
    source_filename: str | None,
) -> tuple[bytes, str]:
    if isinstance(source, (str, Path)):
        path = Path(source)
        return path.read_bytes(), source_filename or path.name
    if isinstance(source, (bytes, bytearray)):
        return bytes(source), source_filename or "weekly-statement.xlsx"
    filename = source_filename or Path(str(getattr(source, "name", ""))).name
    if hasattr(source, "getvalue"):
        return bytes(source.getvalue()), filename or "weekly-statement.xlsx"
    position = source.tell() if hasattr(source, "tell") else None
    data = source.read()
    if position is not None and hasattr(source, "seek"):
        source.seek(position)
    return bytes(data), filename or "weekly-statement.xlsx"


def _read_worksheet_rows(
    workbook_bytes: bytes,
    worksheet: Any,
    required_headers: tuple[str, ...],
    fallback_sheets: list[str],
) -> list[tuple[Any, ...]]:
    rows = list(worksheet.iter_rows(values_only=True))
    declared_abnormal = (worksheet.max_row or 0) <= 1 and (worksheet.max_column or 0) <= 1
    if not declared_abnormal and _find_header(rows, required_headers) is not None:
        return rows
    worksheet_path = _worksheet_archive_path(workbook_bytes, worksheet.title)
    max_row, max_column = _populated_bounds(workbook_bytes, worksheet_path)
    if max_row == 0 or max_column == 0:
        return rows
    worksheet.reset_dimensions()
    fallback_rows = list(worksheet.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_column, values_only=True,
    ))
    if worksheet.title not in fallback_sheets:
        fallback_sheets.append(worksheet.title)
    return fallback_rows

def _worksheet_archive_path(workbook_bytes: bytes, sheet_name: str) -> str:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    doc_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    with ZipFile(BytesIO(workbook_bytes)) as archive:
        workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        rels_xml = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationships = {
        item.attrib["Id"]: item.attrib["Target"]
        for item in rels_xml.findall(f"{{{package_rel_ns}}}Relationship")
    }
    for sheet in workbook_xml.findall(f".//{{{main_ns}}}sheet"):
        if sheet.attrib.get("name") != sheet_name:
            continue
        target = relationships[sheet.attrib[f"{{{doc_rel_ns}}}id"]]
        if target.startswith("/"):
            return target.lstrip("/")
        if target.startswith("xl/"):
            return target
        return posixpath.normpath(posixpath.join("xl", target))
    raise WeeklyStatementParseError(f"Cannot resolve worksheet XML for {sheet_name}.")


def _populated_bounds(workbook_bytes: bytes, worksheet_path: str) -> tuple[int, int]:
    max_row = 0
    max_column = 0
    with ZipFile(BytesIO(workbook_bytes)) as archive:
        with archive.open(worksheet_path) as source:
            for _event, element in ElementTree.iterparse(source, events=("end",)):
                if _local_name(element.tag) != "c":
                    continue
                if _cell_has_populated_value(element):
                    letters, row_number = coordinate_from_string(element.attrib["r"])
                    max_row = max(max_row, row_number)
                    max_column = max(max_column, column_index_from_string(letters))
                element.clear()
    return max_row, max_column


def _cell_has_populated_value(cell: ElementTree.Element) -> bool:
    for child in cell:
        local_name = _local_name(child.tag)
        if local_name in {"v", "f"} and child.text is not None:
            return True
        if local_name == "is" and any(item.text is not None for item in child.iter()):
            return True
    return False


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _find_header(
    rows: list[tuple[Any, ...]], required_headers: tuple[str, ...]
) -> tuple[int, dict[str, int]] | None:
    required = {_normalize_header(header): header for header in required_headers}
    for row_index, row in enumerate(rows):
        positions = {
            _normalize_header(value): column_index
            for column_index, value in enumerate(row)
            if value not in (None, "")
        }
        if required.keys() <= positions.keys():
            return row_index, {
                original: positions[normalized]
                for normalized, original in required.items()
            }
    return None


def _require_header(
    rows: list[tuple[Any, ...]], required_headers: tuple[str, ...], sheet: str
) -> tuple[int, dict[str, int]]:
    found = _find_header(rows, required_headers)
    if found is None:
        raise WeeklyStatementParseError(
            f"{sheet} is missing required column(s): " + ", ".join(required_headers)
        )
    return found


def _find_label_value(rows: list[tuple[Any, ...]], label: str) -> Any:
    normalized = _normalize_header(label)
    for row in rows:
        for index, value in enumerate(row):
            if _normalize_header(value) != normalized:
                continue
            for candidate in row[index + 1:]:
                if candidate not in (None, ""):
                    return candidate
    return None


def _required_label_date(rows: list[tuple[Any, ...]], label: str, sheet: str) -> date:
    parsed = _parse_date(_find_label_value(rows, label))
    if parsed is None:
        raise WeeklyStatementParseError(
            f"Cannot determine statement period {label!r} from {sheet}."
        )
    return parsed


def _required_label_decimal(
    rows: list[tuple[Any, ...]], label: str, sheet: str
) -> Decimal:
    parsed = _parse_decimal(_find_label_value(rows, label))
    if parsed is None:
        raise WeeklyStatementParseError(
            f"Cannot read required control amount {label!r} from {sheet}."
        )
    return parsed


def _optional_label_decimal(rows: list[tuple[Any, ...]], label: str) -> Decimal | None:
    return _parse_decimal(_find_label_value(rows, label))


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _parse_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    text = str(value).strip().replace(",", "")
    if text.casefold().startswith("rm"):
        text = text[2:].strip()
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _text(value: Any) -> str:
    return "" if value in (None, "") else str(value).strip()


def _record_issue(
    issues: list[SourceValueIssue], *, code: str, sheet: str,
    row_number: int, column: str, value: Any,
) -> None:
    issues.append(SourceValueIssue(
        code=code, sheet=sheet, row_number=row_number, column=column,
        message=f"{sheet} row {row_number} has invalid or missing {column}: {value!r}.",
    ))


def _source_mapping(headers: tuple[Any, ...], row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        _text(header): row[index] if index < len(row) else None
        for index, header in enumerate(headers) if _text(header)
    }

def _parse_income_rows(
    rows: list[tuple[Any, ...]], issues: list[SourceValueIssue]
) -> list[SettlementIncomeRow]:
    header_index, positions = _require_header(rows, INCOME_REQUIRED_COLUMNS, "Income")
    headers = rows[header_index]
    parsed: list[SettlementIncomeRow] = []
    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not any(value not in (None, "") for value in row):
            continue
        view_by = _text(row[positions["View By"]])
        order_id = _text(row[positions["Order ID"]])
        order_created = _parse_date(row[positions["Order Creation Date"]])
        payout_completed = _parse_date(row[positions["Payout Completed Date"]])
        total_released = _parse_decimal(row[positions["Total Released Amount (RM)"]])
        components = {
            header: _parse_decimal(row[positions[header]])
            for header in INCOME_COMPONENT_COLUMNS
        }
        if view_by not in {"Order", "Sku"}:
            _record_issue(issues, code="invalid_view_by", sheet="Income", row_number=row_number, column="View By", value=view_by)
        if not order_id:
            _record_issue(issues, code="missing_order_id", sheet="Income", row_number=row_number, column="Order ID", value=order_id)
        if order_created is None:
            _record_issue(issues, code="invalid_order_creation_date", sheet="Income", row_number=row_number, column="Order Creation Date", value=row[positions["Order Creation Date"]])
        if payout_completed is None:
            _record_issue(issues, code="invalid_payout_completed_date", sheet="Income", row_number=row_number, column="Payout Completed Date", value=row[positions["Payout Completed Date"]])
        if total_released is None:
            _record_issue(issues, code="invalid_total_released", sheet="Income", row_number=row_number, column="Total Released Amount (RM)", value=row[positions["Total Released Amount (RM)"]])
        if view_by == "Order":
            for header, value in components.items():
                if value is None:
                    _record_issue(issues, code="invalid_financial_component", sheet="Income", row_number=row_number, column=header, value=row[positions[header]])
        parsed.append(SettlementIncomeRow(
            sequence_no=_text(row[positions["Sequence No."]]),
            view_by=view_by,
            order_id=order_id,
            product_id=_text(row[positions["Product ID"]]),
            product_name=_text(row[positions["Product Name"]]),
            order_creation_date=order_created,
            payout_completed_date=payout_completed,
            release_channel=_text(row[positions["Release Channel"]]),
            order_type=_text(row[positions["Order Type"]]),
            total_released_amount=total_released,
            financial_components=components,
            source_values=_source_mapping(headers, row),
            source_row_number=row_number,
        ))
    return parsed


def _parse_service_fee_rows(
    rows: list[tuple[Any, ...]], issues: list[SourceValueIssue]
) -> list[ServiceFeeDetail]:
    header_index, positions = _require_header(
        rows, SERVICE_FEE_REQUIRED_COLUMNS, "Service Fee Details"
    )
    headers = rows[header_index]
    component_columns = [
        (index, _text(header)) for index, header in enumerate(headers)
        if _text(header) and _text(header) not in SERVICE_FEE_REQUIRED_COLUMNS
    ]
    if not component_columns:
        raise WeeklyStatementParseError(
            "Service Fee Details has no fee breakdown columns."
        )
    parsed: list[ServiceFeeDetail] = []
    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not any(value not in (None, "") for value in row):
            continue
        order_id = _text(row[positions["Order ID"]])
        components = {
            header: _parse_decimal(row[index]) for index, header in component_columns
        }
        if not order_id:
            _record_issue(issues, code="missing_service_fee_order_id", sheet="Service Fee Details", row_number=row_number, column="Order ID", value=order_id)
        for index, header in component_columns:
            if components[header] is None:
                _record_issue(issues, code="invalid_service_fee_component", sheet="Service Fee Details", row_number=row_number, column=header, value=row[index])
        parsed.append(ServiceFeeDetail(
            sequence_no=_text(row[positions["Sequence No."]]),
            order_id=order_id,
            components=components,
            source_row_number=row_number,
        ))
    return parsed

def _parse_shipping_rows(
    rows: list[tuple[Any, ...]], issues: list[SourceValueIssue]
) -> list[ShippingFeeDiscrepancy]:
    header_index, positions = _require_header(
        rows, SHIPPING_REQUIRED_COLUMNS, "Shipping Fee Discrepancy"
    )
    headers = rows[header_index]
    parsed: list[ShippingFeeDiscrepancy] = []
    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not any(value not in (None, "") for value in row):
            continue
        expected = _parse_decimal(row[positions["Expected Shipping Fee:"]])
        actual_header = "Actual Shipping Fee Charged by Logistic Provider:"
        actual = _parse_decimal(row[positions[actual_header]])
        if expected is None:
            _record_issue(issues, code="invalid_expected_shipping_fee", sheet="Shipping Fee Discrepancy", row_number=row_number, column="Expected Shipping Fee:", value=row[positions["Expected Shipping Fee:"]])
        if actual is None:
            _record_issue(issues, code="invalid_actual_shipping_fee", sheet="Shipping Fee Discrepancy", row_number=row_number, column=actual_header, value=row[positions[actual_header]])
        parsed.append(ShippingFeeDiscrepancy(
            order_id=_text(row[positions["Order ID"]]),
            expected_shipping_fee=expected,
            actual_shipping_fee=actual,
            reason=_text(row[positions["Discrepancy reason"]]),
            source_values=_source_mapping(headers, row),
            source_row_number=row_number,
        ))
    return parsed


def _parse_adjustment_rows(
    rows: list[tuple[Any, ...]], issues: list[SourceValueIssue]
) -> list[SettlementAdjustment]:
    header_index, positions = _require_header(
        rows, ADJUSTMENT_REQUIRED_COLUMNS, "Adjustment"
    )
    parsed: list[SettlementAdjustment] = []
    for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if _normalize_header(row[0] if row else None) == _normalize_header("Total Amount"):
            break
        if not any(value not in (None, "") for value in row):
            continue
        amount = _parse_decimal(row[positions["Adjustment Amount"]])
        complete_date = _parse_date(row[positions["Adjustment Complete Date"]])
        payout_date = _parse_date(row[positions["Payout Completed Date"]])
        if amount is None:
            _record_issue(issues, code="invalid_adjustment_amount", sheet="Adjustment", row_number=row_number, column="Adjustment Amount", value=row[positions["Adjustment Amount"]])
        if complete_date is None:
            _record_issue(issues, code="invalid_adjustment_complete_date", sheet="Adjustment", row_number=row_number, column="Adjustment Complete Date", value=row[positions["Adjustment Complete Date"]])
        if payout_date is None:
            _record_issue(issues, code="invalid_adjustment_payout_date", sheet="Adjustment", row_number=row_number, column="Payout Completed Date", value=row[positions["Payout Completed Date"]])
        parsed.append(SettlementAdjustment(
            sequence_no=_text(row[positions["Sequence No."]]),
            adjustment_complete_date=complete_date,
            adjustment_type=_text(row[positions["Adjustment Type | Description"]]),
            adjustment_reason=_text(row[positions["Adjustment Reason"]]),
            adjustment_amount=amount,
            linked_order_id=_text(row[positions["Linked Order No."]]),
            payout_completed_date=payout_date,
            source_row_number=row_number,
        ))
    return parsed